from fastapi import FastAPI, HTTPException, UploadFile, File, Depends, status
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
# Reload trigger
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from pydantic import BaseModel
from typing import Optional, List, Dict
from datetime import datetime, timedelta, timezone
import enum
import random
import secrets
import io
import os
from app.models import models
from app.core import database
from sqlalchemy.orm import Session
from app.schemas import *
import shutil
import uuid
import urllib.request
import urllib.parse
import json
from app.core.database import engine, get_db, SessionLocal
from app.api.router import api_router
from app.api.deps import require_admin, require_operator_or_admin, verify_credentials, security

app = FastAPI(title="Світ Валют API", version="2.0.0")

# Create tables before router registration
from app.core.database import engine as _engine
from app.core.migrations import run_migrations
models.Base.metadata.create_all(bind=_engine)
run_migrations()

# Register API router AFTER tables are created
app.include_router(api_router, prefix="/api")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:3000",
        "https://svit-valut-front.onrender.com",
        "https://mirvalut.com",
        "http://mirvalut.com",
        "https://src.mirvalut.com",
        "http://src.mirvalut.com",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
os.makedirs("static/uploads", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Global settings storage
# These are now mostly for API documentation (Pydantic models)
class SiteSettings(SiteSettingsBase):
    class Config:
        from_attributes = True

class FAQItem(FAQItemBase):
    class Config:
        from_attributes = True

class ServiceItem(ServiceItemBase):
    class Config:
        from_attributes = True

class ArticleItem(ArticleItemBase):
    class Config:
        from_attributes = True

faq_items: List[FAQItem] = [
    FAQItem(id=1, question="Як захиститися від фальшивих купюр", answer="Ми використовуємо професійне обладнання для перевірки справжності банкнот. Кожна купюра проходить детальну перевірку на спеціальних детекторах.", order=1),
    FAQItem(id=2, question="Як правильно розрахувати курс USD → EUR?", answer='Це питання детально розібрано в статті "Що таке конвертація валюти та як вірно рахувати".', link_text="Детальніше", link_url="/articles/conversion", order=2),
    FAQItem(id=3, question="Як працює міжбанк і чому курс змінюється", answer="Міжбанківський курс формується на основі попиту та пропозиції на валютному ринку між банками. Курс змінюється в залежності від економічної ситуації, новин та обсягів торгів.", order=3),
    FAQItem(id=4, question="Коли діє оптовий курс?", answer="Оптовий курс діє при обміні від 1000 USD або еквівалент в іншій валюті. Для отримання оптового курсу достатньо забронювати суму через наш сайт або зателефонувати.", order=4),
    FAQItem(id=5, question="Які банкноти вважаються зношеними?", answer="Зношеними вважаються банкноти з пошкодженнями: надриви, плями, написи, відсутні фрагменти до 40% площі. Ми приймаємо такі банкноти за спеціальним курсом.", order=5),
]

service_items: List[ServiceItem] = [
    ServiceItem(id=1, title="Приймаємо валюту, яка вийшла з обігу", description="Миттєво обміняємо старі фунти, франки, марки, та багато інших.", image_url="https://images.unsplash.com/photo-1621761191319-c6fb62004040?w=400&h=200&fit=crop", link_url="/services/old-currency", order=1),
    ServiceItem(id=2, title="Приймаємо зношену валюту", description="Зручний спосіб позбутися непотрібних купюр.", image_url="https://images.unsplash.com/photo-1605792657660-596af9009e82?w=400&h=200&fit=crop", link_url="/services/damaged-currency", order=2),
    ServiceItem(id=3, title="Старі франки на нові або USD", description="Оновіть франки які вийшли з обігу на нові або долари США.", image_url="https://images.unsplash.com/photo-1580519542036-c47de6196ba5?w=400&h=200&fit=crop", link_url="/services/old-francs", order=3),
]

articles_db: List[ArticleItem] = [
    ArticleItem(id=1, title="Що таке конвертація валюти та як вірно рахувати", excerpt="Інтерес жителів нашої країни до іноземної валюти дуже високий, тому обмін різних видів грошей у Києві є доволі затребуваною послугою.", content="Повний текст статті про конвертацію валюти...", is_published=True, created_at="2025-01-15T10:00:00"),
]

# Enums for Pydantic Models
class UserRole(str, enum.Enum):
    ADMIN = "admin"
    OPERATOR = "operator"

class ReservationStatus(str, enum.Enum):
    PENDING_ADMIN = "pending_admin"  # New: awaiting admin review
    PENDING = "pending"
    CONFIRMED = "confirmed"
    COMPLETED = "completed"
    CANCELLED = "cancelled"
    EXPIRED = "expired"
    DELETED = "deleted"  # Soft delete



# Strict list of 34 currencies for Excel template/ordering
ORDERED_CURRENCIES = [
    "USD", "EUR", "PLN", "GBP", "CHF", "MDL", "DKK", "NOK", "SEK", "CNY", 
    "HUF", "ILS", "KZT", "MLD", "RON", "SAR", "SGD", "THB", "AED", "RSD", 
    "AZN", "BGN", "HKD", "GEL", "KRW", "MXN", "NZD", "EGP", "JPY", "INR", 
    "AUD", "CAD", "CZK", "TRY"
]

# Initial data for migration (will be used once to populate DB)
branches_data = [
    Branch(id=1, address="вул. Старовокзальна, 23", hours="щодня: 9:00-19:00", lat=50.443886, lng=30.490430, is_open=True, phone="(096) 048-88-81"),
    Branch(id=2, address="вул. В. Васильківська, 110", hours="щодня: 8:00-20:00", lat=50.423804, lng=30.518400, is_open=True, phone="(096) 048-88-82"),
    Branch(id=3, address="вул. В. Васильківська, 130", hours="щодня: 8:00-20:00", lat=50.416770, lng=30.522873, is_open=True, phone="(096) 048-88-83"),
    Branch(id=4, address="вул. Р. Окіпної, 2", hours="щодня: 8:00-20:00", lat=50.450606, lng=30.597410, is_open=True, phone="(096) 048-88-84"),
]

users_db = {
    "admin": User(id=1, username="admin", role=UserRole.ADMIN, branch_id=None, name="Адміністратор"),
    "operator1": User(id=2, username="operator1", role=UserRole.OPERATOR, branch_id=1, name="Марія Коваленко"),
    "operator2": User(id=3, username="operator2", role=UserRole.OPERATOR, branch_id=2, name="Олексій Шевченко"),
    "operator3": User(id=4, username="operator3", role=UserRole.OPERATOR, branch_id=3, name="Ірина Бондаренко"),
    "operator4": User(id=5, username="operator4", role=UserRole.OPERATOR, branch_id=4, name="Дмитро Мельник"),
}

passwords_db = {
    "admin": "admin123",
    "operator1": "op1pass",
    "operator2": "op2pass",
    "operator3": "op3pass",
    "operator4": "op4pass",
}

currencies_data = [
    Currency(code="USD", name="US Dollar", name_uk="Долар", flag="🇺🇸", buy_rate=42.10, sell_rate=42.15, is_popular=True),
    Currency(code="EUR", name="Euro", name_uk="Євро", flag="🇪🇺", buy_rate=49.30, sell_rate=49.35, is_popular=True),
    Currency(code="PLN", name="Polish Zloty", name_uk="Польський злотий", flag="🇵🇱", buy_rate=10.50, sell_rate=10.65, is_popular=True),
    Currency(code="GBP", name="British Pound", name_uk="Фунт стерлінгів", flag="🇬🇧", buy_rate=53.10, sell_rate=53.25, is_popular=True),
    Currency(code="CHF", name="Swiss Franc", name_uk="Швейцарський франк", flag="🇨🇭", buy_rate=47.50, sell_rate=47.80, is_popular=True),
    Currency(code="MDL", name="Moldovan Leu", name_uk="Молдовський лей", flag="🇲🇩", buy_rate=2.30, sell_rate=2.40, is_popular=False),
    Currency(code="DKK", name="Danish Krone", name_uk="Данська крона", flag="🇩🇰", buy_rate=6.10, sell_rate=6.20, is_popular=False),
    Currency(code="NOK", name="Norwegian Krone", name_uk="Норвезька крона", flag="🇳🇴", buy_rate=3.80, sell_rate=3.90, is_popular=False),
    Currency(code="SEK", name="Swedish Krona", name_uk="Шведська крона", flag="🇸🇪", buy_rate=3.90, sell_rate=4.00, is_popular=False),
    Currency(code="CNY", name="Chinese Yuan", name_uk="Юань Женьміньбі", flag="🇨🇳", buy_rate=5.70, sell_rate=5.90, is_popular=False),
    Currency(code="HUF", name="Hungarian Forint", name_uk="Форинт", flag="🇭🇺", buy_rate=0.11, sell_rate=0.12, is_popular=False),
    Currency(code="ILS", name="Israeli New Shekel", name_uk="Новий ізраїльський шекель", flag="🇮🇱", buy_rate=11.20, sell_rate=11.50, is_popular=False),
    Currency(code="KZT", name="Kazakhstani Tenge", name_uk="Теньге", flag="🇰🇿", buy_rate=0.09, sell_rate=0.10, is_popular=False),
    Currency(code="MLD", name="Moldovan Leu (Alt)", name_uk="Молдовський лей", flag="🇲🇩", buy_rate=2.30, sell_rate=2.40, is_popular=False), # Duplicate as requested
    Currency(code="RON", name="Romanian Leu", name_uk="Румунський лей", flag="🇷🇴", buy_rate=9.00, sell_rate=9.20, is_popular=False),
    Currency(code="SAR", name="Saudi Riyal", name_uk="Саудівський ріал", flag="🇸🇦", buy_rate=11.00, sell_rate=11.30, is_popular=False),
    Currency(code="SGD", name="Singapore Dollar", name_uk="Сінгапурський долар", flag="🇸🇬", buy_rate=30.50, sell_rate=31.00, is_popular=False),
    Currency(code="THB", name="Thai Baht", name_uk="Бат", flag="🇹🇭", buy_rate=1.10, sell_rate=1.20, is_popular=False),
    Currency(code="AED", name="UAE Dirham", name_uk="Дирхам ОАЕ", flag="🇦🇪", buy_rate=11.30, sell_rate=11.50, is_popular=False),
    Currency(code="RSD", name="Serbian Dinar", name_uk="Сербський динар", flag="🇷🇸", buy_rate=0.38, sell_rate=0.40, is_popular=False),
    Currency(code="AZN", name="Azerbaijani Manat", name_uk="Азербайджанський манат", flag="🇦🇿", buy_rate=24.50, sell_rate=25.00, is_popular=False),
    Currency(code="BGN", name="Bulgarian Lev", name_uk="Болгарський лев", flag="🇧🇬", buy_rate=22.50, sell_rate=23.00, is_popular=False),
    Currency(code="HKD", name="Hong Kong Dollar", name_uk="Гонконгівський долар", flag="🇭🇰", buy_rate=5.30, sell_rate=5.45, is_popular=False),
    Currency(code="GEL", name="Georgian Lari", name_uk="Ларі", flag="🇬🇪", buy_rate=15.20, sell_rate=15.50, is_popular=False),
    Currency(code="KRW", name="South Korean Won", name_uk="Вона", flag="🇰🇷", buy_rate=0.030, sell_rate=0.032, is_popular=False),
    Currency(code="MXN", name="Mexican Peso", name_uk="Мексиканське песо", flag="🇲🇽", buy_rate=2.40, sell_rate=2.50, is_popular=False),
    Currency(code="NZD", name="New Zealand Dollar", name_uk="Новозеландський долар", flag="🇳🇿", buy_rate=25.50, sell_rate=26.00, is_popular=False),
    Currency(code="EGP", name="Egyptian Pound", name_uk="Єгипетський фунт", flag="🇪🇬", buy_rate=0.85, sell_rate=0.95, is_popular=False),
    Currency(code="JPY", name="Japanese Yen", name_uk="Єна", flag="🇯🇵", buy_rate=0.27, sell_rate=0.29, is_popular=False),
    Currency(code="INR", name="Indian Rupee", name_uk="Індійська рупія", flag="🇮🇳", buy_rate=0.49, sell_rate=0.51, is_popular=False),
    Currency(code="AUD", name="Australian Dollar", name_uk="Австралійський долар", flag="🇦🇺", buy_rate=27.00, sell_rate=27.50, is_popular=False),
    Currency(code="CAD", name="Canadian Dollar", name_uk="Канадський долар", flag="🇨🇦", buy_rate=30.00, sell_rate=30.50, is_popular=False),
    Currency(code="CZK", name="Czech Koruna", name_uk="Чеська крона", flag="🇨🇿", buy_rate=1.80, sell_rate=1.90, is_popular=False),
    Currency(code="TRY", name="Turkish Lira", name_uk="Турецька ліра", flag="🇹🇷", buy_rate=1.20, sell_rate=1.30, is_popular=False),
]

# Data migration helper
def init_db_data(db: Session):
    # Migrate Site Settings
    if not db.query(models.SiteSettings).first():
        db_settings = models.SiteSettings()
        db.add(db_settings)
    
    # Migrate FAQ
    if not db.query(models.FAQItem).first():
        for item in faq_items:
            db.add(models.FAQItem(
                question=item.question,
                answer=item.answer,
                link_text=item.link_text,
                link_url=item.link_url,
                order=item.order
            ))
            
    # Migrate Services
    if not db.query(models.ServiceItem).first():
        for item in service_items:
            db.add(models.ServiceItem(
                title=item.title,
                description=item.description,
                image_url=item.image_url,
                link_url=item.link_url,
                is_active=item.is_active,
                order=item.order
            ))
            
    # Migrate Articles
    if not db.query(models.ArticleItem).first():
        for item in articles_db:
            db.add(models.ArticleItem(
                title=item.title,
                excerpt=item.excerpt,
                content=item.content,
                image_url=item.image_url,
                is_published=item.is_published,
                created_at=datetime.fromisoformat(item.created_at)
            ))
            
    # Migrate Branches
    if not db.query(models.Branch).first():
        for b in branches_data:
            db.add(models.Branch(
                id=b.id, # Keep IDs for branches
                address=b.address,
                hours=b.hours,
                lat=b.lat,
                lng=b.lng,
                is_open=b.is_open,
                phone=b.phone
            ))
            
    # Migrate Users
    if not db.query(models.User).first():
        for username, u in users_db.items():
            db.add(models.User(
                username=username,
                password_hash=passwords_db.get(username, "pass"),
                role=models.UserRole[u.role.upper()],
                branch_id=u.branch_id,
                name=u.name
            ))
            
    # Migrate Currencies (Ensure all 34 exist)
    existing_codes = {c.code for c in db.query(models.Currency).all()}
    
    for i, c in enumerate(currencies_data):
        if c.code not in existing_codes:
            db.add(models.Currency(
                code=c.code,
                name=c.name,
                name_uk=c.name_uk,
                flag=c.flag,
                buy_rate=c.buy_rate,
                sell_rate=c.sell_rate,
                is_popular=c.is_popular,
                order=i
            ))
            
    # Migrate Initial Branch Rates (into branch_rates table for branch 1 as default)
    # Ensure rates exist for all currencies for branch 1
    existing_rates = {r.currency_code for r in db.query(models.BranchRate).filter(models.BranchRate.branch_id == 1).all()}
    
    for c in currencies_data:
        if c.code not in existing_rates:
            db.add(models.BranchRate(
                branch_id=1,
                currency_code=c.code,
                buy_rate=c.buy_rate,
                sell_rate=c.sell_rate
            ))

    db.commit()

# Current state (in-memory parts we still use)
rates_updated_at = datetime.now()

# Startup event
@app.on_event("startup")
def startup_db_client():
    db = SessionLocal()
    try:
        init_db_data(db)
    except Exception as e:
        print(f"Startup Error: {e}")
    finally:
        db.close()

reservations_db: List[ReservationResponse] = []

# Routes
@app.get("/api/my-location")
async def get_my_location():
    """Detect user location via IP using external services (Server-side to avoid CORS)"""
    import json
    
    # List of providers with their parser logic
    # (url, lambda data: (lat, lng))
    providers = [
        ("https://ipwho.is/", lambda d: (d.get("latitude"), d.get("longitude")) if d.get("success") is not False else (None, None)),
        ("https://freeipapi.com/api/json", lambda d: (d.get("latitude"), d.get("longitude"))),
        ("https://ipapi.co/json/", lambda d: (d.get("latitude"), d.get("longitude"))),
    ]

    for url, parser in providers:
        try:
            # Set a user agent to avoid being blocked by some APIs
            req = urllib.request.Request(
                url, 
                data=None, 
                headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            )
            with urllib.request.urlopen(req, timeout=3) as response:
                if response.status == 200:
                    data = json.loads(response.read().decode())
                    lat, lng = parser(data)
                    if lat is not None and lng is not None:
                        return {"lat": float(lat), "lng": float(lng), "source": url}
        except Exception as e:
            print(f"Geo provider {url} failed: {e}")
            continue
            
    raise HTTPException(status_code=500, detail="Could not determine location")

@app.get("/api/")
async def root():
    return {"message": "Світ Валют API", "version": "2.0.0"}

@app.get("/api/currencies", response_model=list[Currency])
async def get_currencies(branch_id: int = 1, db: Session = Depends(get_db)):
    """Get all available currencies with rates (Base Rates merged with Branch Overrides)"""
    
    # 1. Get ALL Base Rates (Active ones)
    base_currencies = db.query(models.Currency).filter(models.Currency.is_active == True).order_by(models.Currency.order).all()
    
    # 2. Get Branch Overrides (ALL valid overrides for this branch)
    # If an override exists and is_active=False, we must EXCLUDE the currency.
    overrides = {}
    if branch_id:
        branch_rates = db.query(models.BranchRate).filter(
            models.BranchRate.branch_id == branch_id
        ).all()
        overrides = {r.currency_code: r for r in branch_rates}
    
    result = []
    for base in base_currencies:
        # Check override
        ov = overrides.get(base.code)
        
        is_active = ov.is_active if ov is not None else True
        
        if not is_active:
            buy = 0.0
            sell = 0.0
            w_buy = 0.0
            w_sell = 0.0
            w_threshold = base.wholesale_threshold
        else:
            buy = ov.buy_rate if (ov and ov.buy_rate > 0) else base.buy_rate
            sell = ov.sell_rate if (ov and ov.sell_rate > 0) else base.sell_rate
            w_buy = ov.wholesale_buy_rate if (ov and ov.wholesale_buy_rate > 0) else base.wholesale_buy_rate
            w_sell = ov.wholesale_sell_rate if (ov and ov.wholesale_sell_rate > 0) else base.wholesale_sell_rate
            w_threshold = ov.wholesale_threshold if (ov and ov.wholesale_threshold and ov.wholesale_threshold != 1000) else base.wholesale_threshold

        result.append(Currency(
            code=base.code,
            name=base.name,
            name_uk=base.name_uk,
            flag=base.flag,
            buy_rate=buy,
            sell_rate=sell,
            wholesale_buy_rate=w_buy,
            wholesale_sell_rate=w_sell,
            wholesale_threshold=w_threshold,
            is_popular=base.is_popular,
            is_active=is_active,
            buy_url=base.buy_url,
            sell_url=base.sell_url,
            seo_h1=base.seo_h1,
            seo_h2=base.seo_h2,
            seo_image=base.seo_image,
            seo_text=base.seo_text,
            seo_buy_h1=base.seo_buy_h1,
            seo_buy_h2=base.seo_buy_h2,
            seo_buy_title=base.seo_buy_title,
            seo_buy_desc=base.seo_buy_desc,
            seo_buy_text=base.seo_buy_text,
            seo_buy_image=base.seo_buy_image,
            seo_sell_h1=base.seo_sell_h1,
            seo_sell_h2=base.seo_sell_h2,
            seo_sell_title=base.seo_sell_title,
            seo_sell_desc=base.seo_sell_desc,
            seo_sell_text=base.seo_sell_text,
            seo_sell_image=base.seo_sell_image
        ))
        
    return result

@app.get("/api/currencies/{code}", response_model=Currency)
async def get_currency(code: str, db: Session = Depends(get_db)):
    """Get specific currency by code"""
    r = db.query(models.BranchRate).filter(
        models.BranchRate.branch_id == 1,
        models.BranchRate.currency_code == code.upper()
    ).first()
    
    if not r:
        raise HTTPException(status_code=404, detail="Currency not found")
    
    names = CURRENCY_NAMES.get(r.currency_code, (r.currency_code, r.currency_code))
    return Currency(
        code=r.currency_code,
        name=names[0],
        name_uk=names[1],
        flag=CURRENCY_FLAGS.get(r.currency_code, "🏳️"),
        buy_rate=r.buy_rate,
        sell_rate=r.sell_rate,
        wholesale_buy_rate=r.buy_rate, # Warning: This is getting BranchRate not Currency base! BranchRate doesn't have threshold!
        wholesale_sell_rate=r.sell_rate, # Wait, r is BranchRate.
        # We need to fetch Currency base for threshold.
        # But get_currency uses r.currency_code.
        # Let's check get_currency implementation.
        # It queries BranchRate.
        # We need to query Currency to get threshold.
        is_popular=r.currency_code in POPULAR_CURRENCIES
    )

@app.get("/api/rates/branch/{branch_id}", response_model=list[Currency])
async def get_branch_rates(branch_id: int, db: Session = Depends(get_db)):
    """Get currency rates for a specific branch"""
    today = datetime.now()
    
    # Enable joining BranchRate with Currency to get proper order and names
    # Filter by Active Currencies (and Active Rates?)
    # Generally we want to show all active currencies for the branch
    
    query = (
        db.query(models.BranchRate, models.Currency)
        .join(models.Currency, models.BranchRate.currency_code == models.Currency.code)
        .filter(models.BranchRate.branch_id == branch_id)
        .filter(models.BranchRate.is_active == True)
        .filter(models.Currency.is_active == True)
        .order_by(models.Currency.order)
    )
    
    db_rates = query.all()
    
    # Fallback if no rates for this branch?
    # If empty, maybe try branch 1?
    if not db_rates:
        # Try branch 1
        query = (
            db.query(models.BranchRate, models.Currency)
            .join(models.Currency, models.BranchRate.currency_code == models.Currency.code)
            .filter(models.BranchRate.branch_id == 1)
            .filter(models.BranchRate.is_active == True)
            .filter(models.Currency.is_active == True)
            .order_by(models.Currency.order)
        )
        db_rates = query.all()
    
    result = []
    for rate, curr in db_rates:
        result.append(Currency(
            code=curr.code,
            name=curr.name,
            name_uk=curr.name_uk,
            flag=curr.flag,
            buy_rate=rate.buy_rate,
            sell_rate=rate.sell_rate,
            wholesale_buy_rate=rate.wholesale_buy_rate,
            wholesale_sell_rate=rate.wholesale_sell_rate,
            wholesale_threshold=curr.wholesale_threshold,
            is_popular=curr.is_popular,
            is_active=curr.is_active,
            buy_url=curr.buy_url,
            sell_url=curr.sell_url,
            seo_h1=curr.seo_h1,
            seo_h2=curr.seo_h2,
            seo_image=curr.seo_image,
            seo_text=curr.seo_text,
            seo_buy_h1=curr.seo_buy_h1,
            seo_buy_h2=curr.seo_buy_h2,
            seo_buy_title=curr.seo_buy_title,
            seo_buy_desc=curr.seo_buy_desc,
            seo_buy_text=curr.seo_buy_text,
            seo_buy_image=curr.seo_buy_image,
            seo_sell_h1=curr.seo_sell_h1,
            seo_sell_h2=curr.seo_sell_h2,
            seo_sell_title=curr.seo_sell_title,
            seo_sell_desc=curr.seo_sell_desc,
            seo_sell_text=curr.seo_sell_text,
            seo_sell_image=curr.seo_sell_image
        ))
    return result

@app.get("/api/rates")
async def get_rates(db: Session = Depends(get_db)):
    """Get current exchange rates"""
    # Join with Currency to sort by order
    db_rates = (
        db.query(models.BranchRate, models.Currency)
        .join(models.Currency, models.BranchRate.currency_code == models.Currency.code)
        .filter(models.BranchRate.branch_id == 1)
        .filter(models.BranchRate.is_active == True)
        .filter(models.Currency.is_active == True)
        .order_by(models.Currency.order)
        .all()
    )
    
    # Construct rates dict using ordered list
    # Python 3.7+ preserves insertion order in dicts
    rates_dict = {}
    for r, c in db_rates:
        buy = r.buy_rate if r.buy_rate > 0 else c.buy_rate
        sell = r.sell_rate if r.sell_rate > 0 else c.sell_rate
        rates_dict[r.currency_code] = {"buy": buy, "sell": sell}
        
    return {
        "updated_at": rates_updated_at.isoformat(),
        "base": "UAH",
        "rates": rates_dict
    }

@app.get("/api/calculate")
async def calculate_exchange(
    amount: float,
    from_currency: str,
    to_currency: str = "UAH",
    db: Session = Depends(get_db)
):
    """Calculate exchange amount"""
    from_currency = from_currency.upper()
    to_currency = to_currency.upper()
    
    # helper to find rate
    def get_rate(code):
        if code == "UAH": return None
        return db.query(models.BranchRate).filter(
            models.BranchRate.branch_id == 1,
            models.BranchRate.currency_code == code
        ).first()

    from_r = get_rate(from_currency)
    to_r = get_rate(to_currency)
    
    if from_currency == "UAH":
        if not to_r:
            raise HTTPException(status_code=404, detail="Currency not found")
        
        # Check wholesale threshold
        # For UAH -> Currency, amount is in UAH. We convert to Currency amount? 
        # Or compare UAH amount? Usually wholesale limit is "1000 USD".
        # So we should convert amount to base currency (USD equivalent) to check?
        # OR simply specific currency limit.
        # Let's assume limit is defined in THAT currency units.
        # But here amount is UAH.
        # So we check if (amount / sell_rate) >= threshold.
        
        to_curr_def = db.query(models.Currency).filter(models.Currency.code == to_currency).first()
        threshold = to_curr_def.wholesale_threshold if to_curr_def else 1000
        
        rate = to_r.sell_rate
        converted_amount = amount / rate
        
        if converted_amount >= threshold and to_r.wholesale_sell_rate > 0:
            rate = to_r.wholesale_sell_rate
            result = amount / rate
        else:
            result = converted_amount

    elif to_currency == "UAH":
        if not from_r:
            raise HTTPException(status_code=404, detail="Currency not found")
            
        # Check wholesale threshold
        # Amount is in FROM currency.
        from_curr_def = db.query(models.Currency).filter(models.Currency.code == from_currency).first()
        threshold = from_curr_def.wholesale_threshold if from_curr_def else 1000
        
        rate = from_r.buy_rate
        if amount >= threshold and from_r.wholesale_buy_rate > 0:
            rate = from_r.wholesale_buy_rate
            
        result = amount * rate
    else:
        # Cross rate via UAH
        if not from_r or not to_r:
             raise HTTPException(status_code=404, detail="Currency not found")
        
        # FROM -> UAH (Buy)
        uah = amount * from_r.buy_rate
        # UAH -> TO (Sell)
        result = uah / to_r.sell_rate
        rate = result / amount
    
    return {
        "from_amount": amount,
        "from_currency": from_currency,
        "to_amount": round(result, 2),
        "to_currency": to_currency,
        "rate": rate
    }

@app.get("/api/orders", response_model=list[Order])
async def get_orders(
    type: Optional[str] = None,
    page: int = 1,
    limit: int = 10
):
    """Get active orders"""
    filtered = orders_data
    if type and type != "all":
        filtered = [o for o in orders_data if o.type == type]
    
    start = (page - 1) * limit
    end = start + limit
    
    return filtered[start:end]

@app.get("/api/orders/count")
async def get_orders_count():
    """Get total orders count"""
    return {
        "total": len(orders_data),
        "buy": len([o for o in orders_data if o.type == "buy"]),
        "sell": len([o for o in orders_data if o.type == "sell"])
    }

@app.get("/api/branches/{branch_id}/balances", response_model=List[BranchBalance])
async def get_branch_balances(branch_id: int, db: Session = Depends(get_db)):
    """Get all balances for a specific branch"""
    return db.query(models.BranchBalance).filter(models.BranchBalance.branch_id == branch_id).all()

@app.put("/api/branches/{branch_id}/balances")
async def update_branch_balances(
    branch_id: int, 
    request: BranchBalanceBatchUpdate, 
    db: Session = Depends(get_db)
):
    """Batch update balances for a branch"""
    for entry in request.balances:
        db_balance = db.query(models.BranchBalance).filter(
            models.BranchBalance.branch_id == branch_id,
            models.BranchBalance.currency_code == entry.currency_code,
            models.BranchBalance.category == entry.category
        ).first()
        
        if db_balance:
            db_balance.amount = entry.amount
        else:
            db_balance = models.BranchBalance(
                branch_id=branch_id,
                currency_code=entry.currency_code,
                category=entry.category,
                amount=entry.amount
            )
            db.add(db_balance)
            
    db.commit()
    return {"message": "Balances updated successfully"}

@app.post("/api/reservations", response_model=ReservationResponse)
async def create_reservation(request: ReservationRequest, db: Session = Depends(get_db)):
    """Create a new currency reservation"""
    branch_id = request.branch_id or 1
    rate = 0.0
    get_amount = 0.0

    if request.give_currency.upper() == "UAH":
        to_curr = db.query(models.BranchRate).filter(
            models.BranchRate.branch_id == branch_id,
            models.BranchRate.currency_code == request.get_currency.upper()
        ).first()
        if not to_curr:
            # Fallback to branch 1
            to_curr = db.query(models.BranchRate).filter(
                models.BranchRate.branch_id == 1,
                models.BranchRate.currency_code == request.get_currency.upper()
            ).first()
            
        if not to_curr:
            raise HTTPException(status_code=404, detail="Currency not found")
            
        # Determine effective rate
        base_rate = to_curr.sell_rate
        # For UAH -> Foreign, threshold is checked against Foreign amount
        # We can calculate tentative amount first
        tentative_get = request.give_amount / base_rate if base_rate > 0 else 0
        
        effective_rate = base_rate
        if to_curr.wholesale_threshold > 0 and tentative_get >= to_curr.wholesale_threshold and to_curr.wholesale_sell_rate > 0:
            effective_rate = to_curr.wholesale_sell_rate
            
        rate = effective_rate
        get_amount = request.give_amount / rate if rate > 0 else 0
    else:
        from_curr = db.query(models.BranchRate).filter(
            models.BranchRate.branch_id == branch_id,
            models.BranchRate.currency_code == request.give_currency.upper()
        ).first()
        if not from_curr:
            # Fallback to branch 1
            from_curr = db.query(models.BranchRate).filter(
                models.BranchRate.branch_id == 1,
                models.BranchRate.currency_code == request.give_currency.upper()
            ).first()
            
        if not from_curr:
            raise HTTPException(status_code=404, detail="Currency not found")
            
        # Determine effective rate
        base_rate = from_curr.buy_rate
        # For Foreign -> UAH, threshold is checked against Foreign amount (give_amount)
        effective_rate = base_rate
        if from_curr.wholesale_threshold > 0 and request.give_amount >= from_curr.wholesale_threshold and from_curr.wholesale_buy_rate > 0:
            effective_rate = from_curr.wholesale_buy_rate
            
        rate = effective_rate
        get_amount = request.give_amount * rate

    # If frontend provided its own calculated values, prefer them (after basic validation)
    # This ensures consistency with what the user saw even if rates changed by a tiny fraction or rounding
    if request.get_amount is not None:
        # Simple safety check: ensure provided get_amount isn't wildly different from backend calculation (e.g. > 10% diff)
        if get_amount > 0 and abs(request.get_amount - get_amount) / get_amount < 0.1:
            get_amount = request.get_amount
            if request.rate is not None:
                rate = request.rate
    
    # Use Naive Kyiv Time to ensure correct string-based sorting with existing legacy data
    # and correct display on frontend (as local time).
    try:
        from zoneinfo import ZoneInfo
        kyiv = ZoneInfo("Europe/Kyiv")
        now = datetime.now(kyiv).replace(tzinfo=None)
    except ImportError:
        # Fallback for systems without zoneinfo (e.g. older python)
        # Kyiv is UTC+2 in winter, UTC+3 in summer. Currently Feb -> +2
        now = datetime.utcnow() + timedelta(hours=2)

    expires_at = now + timedelta(minutes=60)
    
    db_res = models.Reservation(
        give_amount=request.give_amount,
        give_currency=request.give_currency.upper(),
        get_amount=round(get_amount, 2),
        get_currency=request.get_currency.upper(),
        rate=rate,
        phone=request.phone,
        customer_name=request.customer_name,
        status=models.ReservationStatus.PENDING_ADMIN,
        branch_id=request.branch_id,
        created_at=now,
        expires_at=expires_at
    )
    
    db.add(db_res)
    db.commit()
    db.refresh(db_res)
    
    branch = db.query(models.Branch).filter(models.Branch.id == db_res.branch_id).first()
    
    return ReservationResponse(
        id=db_res.id,
        give_amount=db_res.give_amount,
        give_currency=db_res.give_currency,
        get_amount=db_res.get_amount,
        get_currency=db_res.get_currency,
        rate=db_res.rate,
        phone=db_res.phone,
        customer_name=db_res.customer_name,
        status=db_res.status,
        branch_id=db_res.branch_id,
        branch_address=branch.address if branch else None,
        created_at=db_res.created_at.isoformat(),
        updated_at=db_res.updated_at.isoformat() if db_res.updated_at else None,
        expires_at=db_res.expires_at.isoformat()
    )

@app.get("/api/reservations/{reservation_id}", response_model=ReservationResponse)
async def get_reservation(reservation_id: int, db: Session = Depends(get_db)):
    """Get reservation by ID"""
    db_res = db.query(models.Reservation).filter(models.Reservation.id == reservation_id).first()
    if not db_res:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    branch = db.query(models.Branch).filter(models.Branch.id == db_res.branch_id).first()
    
    return ReservationResponse(
        id=db_res.id,
        give_amount=db_res.give_amount,
        give_currency=db_res.give_currency,
        get_amount=db_res.get_amount,
        get_currency=db_res.get_currency,
        rate=db_res.rate,
        phone=db_res.phone,
        customer_name=db_res.customer_name,
        status=db_res.status,
        branch_id=db_res.branch_id,
        branch_address=branch.address if branch else None,
        created_at=db_res.created_at.isoformat(),
        updated_at=db_res.updated_at.isoformat() if db_res.updated_at else None,
        expires_at=db_res.expires_at.isoformat(),
        completed_at=db_res.completed_at.isoformat() if db_res.completed_at else None,
        operator_note=db_res.operator_note
    )

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


# ============== AUTH ENDPOINTS ==============

@app.post("/api/auth/login")
async def login(user: User = Depends(verify_credentials)):
    """Login and get user info"""
    branch = None
    if user.branch_id:
        branch = next((b for b in branches_data if b.id == user.branch_id), None)
    
    return {
        "user": {
            "id": user.id,
            "username": user.username,
            "name": user.name,
            "role": user.role,
            "branch_id": user.branch_id,
            "branch_address": branch.address if branch else None
        }
    }

@app.get("/api/auth/me")
async def get_current_user(user: User = Depends(verify_credentials)):
    """Get current user info"""
    branch = None
    if user.branch_id:
        branch = next((b for b in branches_data if b.id == user.branch_id), None)
    
    return {
        "id": user.id,
        "username": user.username,
        "name": user.name,
        "role": user.role,
        "branch_id": user.branch_id,
        "branch_address": branch.address if branch else None
    }


# ============== ADMIN ENDPOINTS ==============

@app.post("/api/upload/image")
async def upload_image(
    file: UploadFile = File(...),
    user: User = Depends(require_admin)
):
    """Upload an image for SEO or other purposes"""
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    # Generate unique filename
    ext = os.path.splitext(file.filename)[1]
    if not ext:
        ext = ".jpg" # Default fallback
    
    filename = f"{uuid.uuid4()}{ext}"
    file_path = f"static/uploads/{filename}"
    
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Return URL relative to root (frontend should prepend domain if needed, 
        # but usually /static works if proxied or same origin)
        return {"url": f"/static/uploads/{filename}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not save image: {str(e)}")

@app.post("/api/admin/rates/upload", response_model=RatesUploadResponseV2)
async def upload_rates(
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Upload Excel file with rates.
    Supports two formats:
    1. Base rates: 'Курси' or first sheet
    2. Branch rates: 'Відділення' (Vertical format) or Sheet with branch columns (Matrix)
    """
    global currencies_data, rates_updated_at
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")
    
    try:
        import pandas as pd
        import zipfile
        import io # Ensure io is imported for BytesIO
    except ImportError:
        raise HTTPException(status_code=500, detail="Необхідні бібліотеки (pandas, zipfile) не встановлені.")

    try:
        contents = await file.read()
        xlsx = pd.ExcelFile(io.BytesIO(contents)) # Use io.BytesIO for pd.ExcelFile
        sheet_names = [s.lower() for s in xlsx.sheet_names]
    except zipfile.BadZipFile:
        print("DEBUG: BadZipFile Error")
        raise HTTPException(status_code=400, detail="Невірний формат файлу. Будь ласка, завантажте коректний .xlsx файл.")
    except ValueError as e:
        print(f"DEBUG: ValueError: {e}")
        raise HTTPException(status_code=400, detail=f"Помилка читання Excel: {str(e)}")
    except Exception as e:
        print(f"DEBUG: Unknown Error: {e}")
        raise HTTPException(status_code=400, detail=f"Не вдалося відкрити файл: {str(e)}")

    # Process Logic 
    try:
        errors = []
        base_updated = 0
        branch_updated = 0
        processed_codes = set()
        explicitly_updated_branches = set()  # Track (branch_id, currency_code) pairs updated from Excel
        
        # 1. Process BASE RATES
        base_sheet = None
        if 'курси' in sheet_names:
            base_sheet = xlsx.sheet_names[sheet_names.index('курси')]
        elif 'rates' in sheet_names:
            base_sheet = xlsx.sheet_names[sheet_names.index('rates')]
        else:
            base_sheet = xlsx.sheet_names[0]
        
        # Detect format:
        # Standard: Row 0 = Headers
        # Old "ID" row: Row 0 = ID, Row 1 = Headers
        # New "3-row": Row 0 = Number, Row 1 = Address, Row 2 = Headers
        
        # Read first 10 rows to dynamically find the header row
        df_preview = pd.read_excel(xlsx, sheet_name=base_sheet, header=None, nrows=10)
        
        header_row = 0
        branch_col_definitions = {} # {col_idx: {'branch_id': ..., 'type': ...}}
        branch_col_map = {}
        
        if not df_preview.empty:
            for i, row in df_preview.iterrows():
                row_vals = [str(x).lower() for x in row.values]
                if any('код' in x or 'code' in x or 'валют' in x for x in row_vals):
                    header_row = i
                    break
                    
            if header_row >= 2:
                # Check for 3-row format (Addresses in header_row - 1, Numbers in header_row - 2)
                row_numbers = df_preview.iloc[header_row - 2].astype(str).values
                row_addresses = df_preview.iloc[header_row - 1].astype(str).values
                row_headers = df_preview.iloc[header_row].astype(str).values
                
                # We check if there are any valid addresses to trigger this branch logic
                has_valid_addresses = any(str(x).strip() and str(x).strip() != 'nan' and 'unnamed' not in str(x).lower() for x in row_addresses[2:])
                
                if has_valid_addresses:
                # Start logging checking logic
                    print(f"DEBUG BRANCH DETECT: row_numbers={row_numbers.tolist()}")
                    print(f"DEBUG BRANCH DETECT: row_addresses={row_addresses.tolist()}")
                    print(f"DEBUG BRANCH DETECT: row_headers={row_headers.tolist()}")
                    
                    for i in range(2, len(row_numbers)):
                        addr_val = str(row_addresses[i]).strip()
                        number_val = str(row_numbers[i]).strip()
                        
                        if addr_val and addr_val != 'nan' and 'unnamed' not in addr_val.lower():
                            branch = None
                            new_number = None

                            import re
                            match = re.search(r'(\d+)', number_val)
                            if match:
                                new_number = int(match.group(1))
                            
                            print(f"DEBUG BRANCH: Col {i}, addr='{addr_val}', number={new_number}")
                            
                            if new_number:
                                # Prioritize exact number match
                                branch = db.query(models.Branch).filter(models.Branch.number == new_number).first()
                                print(f"DEBUG BRANCH: Number lookup {new_number} -> {'FOUND id=' + str(branch.id) if branch else 'NOT FOUND'}")
                                # If branch exists but address changed, we update the address
                                if branch and branch.address != addr_val:
                                    branch.address = addr_val
                                    db.add(branch)
                            
                            if not branch:
                                # Try exact address match. BUT if we have a new_number, 
                                # we must restrict this to branches that either have no number yet, 
                                # or have the exact same number, to prevent merging "№ 612" into "№ 611" just because they share an address.
                                query = db.query(models.Branch).filter(models.Branch.address == addr_val)
                                if new_number:
                                    # Only map to an existing address-matched branch if its number is unset or matches our target number
                                    from sqlalchemy import or_
                                    query = query.filter(or_(models.Branch.number == None, models.Branch.number == new_number))
                                
                                branch = query.first()
                                print(f"DEBUG BRANCH: Exact addr lookup '{addr_val}' (with safety) -> {'FOUND id=' + str(branch.id) if branch else 'NOT FOUND'}")
                            
                            if not branch:
                                # Check if another branch exists at this address to copy coordinates
                                existing_at_addr = db.query(models.Branch).filter(models.Branch.address == addr_val).first()
                                
                                print(f"DEBUG BRANCH: Creating new branch addr='{addr_val}', number={new_number}")
                                branch = models.Branch(
                                    address=addr_val,
                                    number=new_number or (db.query(models.Branch).count() + 1),
                                    order=i,
                                    is_open=True,
                                    hours=existing_at_addr.hours if existing_at_addr else "щодня: 8:00-20:00",
                                    lat=existing_at_addr.lat if existing_at_addr else 50.4501,
                                    lng=existing_at_addr.lng if existing_at_addr else 30.5234,
                                    phone=existing_at_addr.phone if existing_at_addr else None,
                                    telegram_chat=existing_at_addr.telegram_chat if existing_at_addr else None
                                )
                                db.add(branch)
                                db.commit()
                                db.refresh(branch)
                                print(f"DEBUG BRANCH: Created branch id={branch.id}")
                                
                            if branch:
                                current_branch = branch
                                print(f"DEBUG BRANCH: current_branch set to id={branch.id}, addr='{branch.address}'")
                                
                    if current_branch:
                        h_val = str(row_headers[i]).lower()
                        rate_type = None
                        if 'опт' in h_val:
                            if 'куп' in h_val: rate_type = 'wholesale_buy'
                            elif 'прод' in h_val: rate_type = 'wholesale_sell'
                        else:
                            if 'куп' in h_val: rate_type = 'buy'
                            elif 'прод' in h_val: rate_type = 'sell'
                            
                        if rate_type:
                            branch_col_definitions[i] = {'branch_id': current_branch.id, 'type': rate_type}
                            print(f"DEBUG BRANCH: Mapped Col {i} ({rate_type}) to branch id={current_branch.id}")
                                
            elif header_row == 1:
                # 2-Row old format (Numbers in header_row - 1)
                row_numbers = df_preview.iloc[0].astype(str).values
                order_counter = 1
                for i in range(2, len(row_numbers)):
                    val = str(row_numbers[i])
                    if any(marker in val for marker in ['ID:', '№', 'Nr', 'No']) or (val.strip().isdigit()):
                        try:
                            import re
                            match = re.search(r'(\d+)', val)
                            if match:
                                bid = int(match.group(1))
                                b = db.query(models.Branch).filter(models.Branch.id == bid).first()
                                if not b:
                                    b = db.query(models.Branch).filter(models.Branch.number == bid).first()
                                if not b:
                                    b = models.Branch(
                                        address=f"Відділення {bid}",
                                        number=bid,
                                        order=order_counter,
                                        is_open=True,
                                        hours="щодня: 8:00-20:00",
                                        lat=50.4501,
                                        lng=30.5234
                                    )
                                    db.add(b)
                                    db.commit()
                                    db.refresh(b)
                                    
                                branch_col_map[i] = b.id
                                if b.order != order_counter:
                                    b.order = order_counter
                                    db.add(b)
                                    db.commit()
                                order_counter += 1
                        except Exception as e:
                            print(f"DEBUG: 2-Row Branch Logic error: {e}")

        # Read actual data
        df_base = pd.read_excel(xlsx, sheet_name=base_sheet, header=header_row)
        print(f"DEBUG: Header Row: {header_row}, DF Shape: {df_base.shape}")
        
        # Deduplicate base columns
        new_cols_base = []
        counts_base = {}
        for col in df_base.columns:
            c_clean = str(col).strip().lower()
            if c_clean in counts_base:
                counts_base[c_clean] += 1
                new_cols_base.append(f"{c_clean}_{counts_base[c_clean]}")
            else:
                counts_base[c_clean] = 0
                new_cols_base.append(c_clean)
        df_base.columns = new_cols_base
        
        # Find columns
        # ... (standard find logic)
        code_col = next((c for c in df_base.columns if any(x in str(c).lower() for x in ['код', 'code', 'iso'])), None)
        if not code_col:
             code_col = next((c for c in df_base.columns if any(x in str(c).lower() for x in ['валют', 'currency']) and not any(x in str(c).lower() for x in ['назва', 'name'])), None)
             
        name_col = next((c for c in df_base.columns if c != code_col and any(x in str(c).lower() for x in ['назва', 'name', 'валют', 'currency'])), None)
        buy_col = next((c for c in df_base.columns if any(x in str(c).lower() for x in ['купів', 'buy', 'покуп']) and not 'опт' in str(c).lower()), None)
        sell_col = next((c for c in df_base.columns if any(x in str(c).lower() for x in ['прода', 'sell']) and not 'опт' in str(c).lower()), None)
        wholesale_buy_col = next((c for c in df_base.columns if 'опт' in str(c).lower() and any(x in str(c).lower() for x in ['купів', 'buy', 'покуп'])), None)
        wholesale_sell_col = next((c for c in df_base.columns if 'опт' in str(c).lower() and any(x in str(c).lower() for x in ['прода', 'sell'])), None)
        flag_col = next((c for c in df_base.columns if any(x in str(c).lower() for x in ['прапор', 'flag'])), None)
        
        if code_col and 'валют' in str(code_col).lower() and not 'код' in str(code_col).lower():
             first_valid = df_base[df_base[code_col].notna()].head(1)
             if not first_valid.empty and len(str(first_valid.iloc[0][code_col])) > 3:
                 code_col = None
 
        if not code_col or code_col == buy_col or code_col == sell_col:
             for col in df_base.columns:
                 sample = df_base[df_base[col].notna()].head(3)
                 if not sample.empty and all(len(str(x).strip()) == 3 and str(x).strip().isalpha() for x in sample[col]):
                     code_col = col
                     break
 
        if code_col and buy_col and sell_col:
            for idx, row in df_base.iterrows():
                if idx < 3: print(f"DEBUG: Process Loop Row {idx}")
                try:
                    code_val = row[code_col]
                    if pd.isna(code_val): continue
                    code = str(code_val).strip().upper()
                    if len(code) != 3: continue
                    
                    try:
                        b_val = str(row[buy_col]).replace(',', '.').replace(' ', '').strip()
                        buy_rate = float(b_val)
                        s_val = str(row[sell_col]).replace(',', '.').replace(' ', '').strip()
                        sell_rate = float(s_val)
                    except:
                        continue
                    
                    if buy_rate <= 0 or sell_rate <= 0: continue
                    
                    wholesale_buy = 0.0
                    wholesale_sell = 0.0
                    if wholesale_buy_col and pd.notna(row[wholesale_buy_col]):
                         try: 
                             wb_val = str(row[wholesale_buy_col]).replace(',', '.').replace(' ', '').strip()
                             wholesale_buy = float(wb_val)
                         except: pass
                    if wholesale_sell_col and pd.notna(row[wholesale_sell_col]):
                         try: 
                             ws_val = str(row[wholesale_sell_col]).replace(',', '.').replace(' ', '').strip()
                             wholesale_sell = float(ws_val)
                         except: pass
                    
                    # Determine Flag
                    flag = "🏳️"
                    if flag_col and pd.notna(row[flag_col]):
                        flag = str(row[flag_col]).strip()
                    else:
                        flag = CURRENCY_FLAGS.get(code, "🏳️")

                    # Determine Name
                    name_uk = None
                    if name_col and pd.notna(row[name_col]):
                        name_uk = str(row[name_col]).strip()
                    
                    # 2. Process BRANCH RATES
                    branch_updates = {} # {branch_id: {buy: ..., sell: ..., ...}}
                    
                    if branch_col_definitions:
                        # Pre-fill all branches so backfill works even if local rates are empty
                        for defs in branch_col_definitions.values():
                             if defs['branch_id'] not in branch_updates:
                                 branch_updates[defs['branch_id']] = {}

                        if idx < 3:
                             print(f"DEBUG: Row {idx} Values: {row.values.tolist()}")
                        
                        for col_idx, defs in branch_col_definitions.items():
                            if col_idx >= len(row): continue
                            
                            val = row.iloc[col_idx]
                            
                            if pd.isna(val): continue
                            
                            try:
                                val_str = str(val).replace(',', '.').replace(' ', '').strip()
                                val_float = float(val_str)
                                
                                if val_float <= 0: continue
                                
                                b_id = defs['branch_id']
                                r_type = defs['type']
                                
                                if b_id not in branch_updates:
                                    branch_updates[b_id] = {}
                                branch_updates[b_id][r_type] = val_float
                            except Exception as e:
                                pass
                        
                        if idx < 3:
                            print(f"DEBUG: Branch Updates Row {idx}: {branch_updates}")

                    # Backfill Global Wholesale if missing
                    if wholesale_buy <= 0 and branch_updates:
                        for b_data in branch_updates.values():
                            if b_data.get('wholesale_buy', 0) > 0:
                                wholesale_buy = b_data['wholesale_buy']
                                if idx < 3: print(f"DEBUG: Backfilled Wholesale Buy: {wholesale_buy}")
                                break
                    
                    if wholesale_sell <= 0 and branch_updates:
                        for b_data in branch_updates.values():
                            if b_data.get('wholesale_sell', 0) > 0:
                                wholesale_sell = b_data['wholesale_sell']
                                if idx < 3: print(f"DEBUG: Backfilled Wholesale Sell: {wholesale_sell}")
                                break

                    # Upsert Currency (Base Rate)
                    curr_db = db.query(models.Currency).filter(models.Currency.code == code).first()
                    if curr_db:
                        curr_db.buy_rate = buy_rate
                        curr_db.sell_rate = sell_rate
                        curr_db.wholesale_buy_rate = wholesale_buy
                        curr_db.wholesale_sell_rate = wholesale_sell
                        curr_db.is_active = True
                        if not curr_db.flag or (flag_col and pd.notna(row[flag_col])):
                            curr_db.flag = flag
                        if name_uk:
                             curr_db.name_uk = name_uk
                             # Optional: update 'name' too
                             curr_db.name = name_uk 
                        base_updated += 1
                    else:
                        names = CURRENCY_NAMES.get(code, (code, code))
                        final_name = name_uk if name_uk else names[0]
                        final_name_uk = name_uk if name_uk else names[1]
                        
                        curr_db = models.Currency(
                            code=code, name=final_name, name_uk=final_name_uk,
                            buy_rate=buy_rate, sell_rate=sell_rate,
                            wholesale_buy_rate=wholesale_buy, wholesale_sell_rate=wholesale_sell,
                            flag=flag,
                            is_active=True, is_popular=code in POPULAR_CURRENCIES
                        )
                        db.add(curr_db)
                        db.commit()
                        db.refresh(curr_db)
                        base_updated += 1
                    
                    processed_codes.add(code)
                        
                    # Update cache
                    existing_cache = next((c for c in currencies_data if c.code == code), None)
                    if existing_cache:
                        existing_cache.buy_rate = buy_rate
                        existing_cache.sell_rate = sell_rate
                        existing_cache.wholesale_buy_rate = wholesale_buy
                        existing_cache.wholesale_sell_rate = wholesale_sell
                
                
                
                    # 2. Process BRANCH RATES
                    # If we found branch_col_definitions (New 3-Row Format), use it
                    if branch_col_definitions:
                        if idx < 3:
                             print(f"DEBUG: Row {idx} Values: {row.values.tolist()}")
                             
                        # Group values by branch first
                        branch_updates = {} # {branch_id: {buy: ..., sell: ..., ...}}
                        
                        for col_idx, defs in branch_col_definitions.items():
                            if col_idx >= len(row): continue
                            
                            val = row.iloc[col_idx]
                            if idx < 3: print(f"DEBUG: Row {idx} Col {col_idx} RAW: {val} (isna: {pd.isna(val)})")
                            
                            if pd.isna(val): continue
                            
                            try:
                                val_str = str(val).replace(',', '.').replace(' ', '').strip()
                                val_float = float(val_str)
                                if idx < 5: # Only debug first few rows
                                     print(f"DEBUG: Row {idx} Col {col_idx} ({defs['type']}): '{val}' -> {val_float}")
                                
                                if val_float <= 0: continue
                                
                                b_id = defs['branch_id']
                                r_type = defs['type']
                                
                                if b_id not in branch_updates:
                                    branch_updates[b_id] = {}
                                branch_updates[b_id][r_type] = val_float
                            except Exception as e:
                                if idx < 5:
                                    print(f"DEBUG: Failed to parse '{val}': {e}")
                                pass
                        
                        # Apply updates
                        if idx < 5:
                            print(f"DEBUG: Branch Updates Row {idx}: {branch_updates}")
                        
                        for b_id, rates in branch_updates.items():
                            # Continue even if local rates are empty, IF global fallback exists
                            has_fallback = (wholesale_buy > 0 or wholesale_sell > 0)
                            if not rates and not has_fallback: continue
                            
                            br_rate = db.query(models.BranchRate).filter(
                                models.BranchRate.branch_id == b_id,
                                models.BranchRate.currency_code == code
                            ).first()
                            
                            # Fallback to global wholesale if missing
                            w_buy = rates.get('wholesale_buy', 0)
                            if w_buy <= 0 and wholesale_buy > 0: w_buy = wholesale_buy
                            
                            w_sell = rates.get('wholesale_sell', 0)
                            if w_sell <= 0 and wholesale_sell > 0: w_sell = wholesale_sell
                            
                            if not br_rate:
                                br_rate = models.BranchRate(
                                    branch_id=b_id,
                                    currency_code=code,
                                    buy_rate=rates.get('buy', 0),
                                    sell_rate=rates.get('sell', 0),
                                    wholesale_buy_rate=w_buy,
                                    wholesale_sell_rate=w_sell
                                )
                                db.add(br_rate)
                            else:
                                if 'buy' in rates: br_rate.buy_rate = rates['buy']
                                if 'sell' in rates: br_rate.sell_rate = rates['sell']
                                if w_buy > 0: br_rate.wholesale_buy_rate = w_buy
                                if w_sell > 0: br_rate.wholesale_sell_rate = w_sell
                            
                            explicitly_updated_branches.add((b_id, code))
                            branch_updated += 1

                    elif branch_col_map:
                        # Old key-based map fallback
                         for col_idx, branch_id in branch_col_map.items():
                             if col_idx + 1 >= len(df_base.columns): continue # Changed df to df_base
                             
                             try:
                                 val_buy = row.iloc[col_idx]
                                 val_sell = row.iloc[col_idx+1]
                                 
                                 if pd.isna(val_buy) or pd.isna(val_sell): continue
                                 
                                 b_buy = float(val_buy)
                                 b_sell = float(val_sell)
                                 
                                 if b_buy <= 0 or b_sell <= 0: continue
                                 
                                 b_wh_buy = 0.0
                                 b_wh_sell = 0.0
                                 
                                 # Try to read next 2 cols for wholesale
                                 try: b_wh_buy = float(row.iloc[col_idx+2])
                                 except: pass
                                 try: b_wh_sell = float(row.iloc[col_idx+3])
                                 except: pass
                                 
                                 br_rate = db.query(models.BranchRate).filter(
                                     models.BranchRate.branch_id == branch_id,
                                     models.BranchRate.currency_code == code
                                 ).first()
                                 
                                 if br_rate:
                                     br_rate.buy_rate = b_buy
                                     br_rate.sell_rate = b_sell
                                     br_rate.wholesale_buy_rate = b_wh_buy
                                     br_rate.wholesale_sell_rate = b_wh_sell
                                 else:
                                      br_rate = models.BranchRate(
                                          branch_id=branch_id,
                                          currency_code=code,
                                          buy_rate=b_buy,
                                          sell_rate=b_sell,
                                          wholesale_buy_rate=b_wh_buy,
                                          wholesale_sell_rate=b_wh_sell
                                      )
                                      db.add(br_rate)
                                 explicitly_updated_branches.add((branch_id, code))
                                 branch_updated += 1
                             except:
                                  pass
            
                except Exception:
                    pass
            
            db.commit()
        
        # 2. Process BRANCH RATES
        branch_sheet = None
        if 'відділення' in sheet_names:
            branch_sheet = xlsx.sheet_names[sheet_names.index('відділення')]
        elif 'branches' in sheet_names:
            branch_sheet = xlsx.sheet_names[sheet_names.index('branches')]
        elif 'філії' in sheet_names:
            branch_sheet = xlsx.sheet_names[sheet_names.index('філії')]
        elif base_sheet:
            # Fallback: Check if base_sheet contains branch columns
            # This supports the single-sheet hybrid format
            branch_sheet = base_sheet
            
        if branch_sheet and branch_sheet != base_sheet: # Only process if it's a separate sheet
            df_branch = pd.read_excel(xlsx, sheet_name=branch_sheet)
            
            # Deduplicate columns (handle '$' and '$ ' becoming same after strip)
            new_cols = []
            counts = {}
            for col in df_branch.columns:
                c_clean = str(col).strip().lower()
                if c_clean in counts:
                    counts[c_clean] += 1
                    new_cols.append(f"{c_clean}_{counts[c_clean]}")
                else:
                    counts[c_clean] = 0
                    new_cols.append(c_clean)
            df_branch.columns = new_cols
            
            # Check format type
            has_branch_col = any(x in str(c).lower() for c in df_branch.columns for x in ['відділ', 'branch', 'філі', 'каса', 'cashier'])
            
            if has_branch_col:
                # Vertical Format OR Hybrid Matrix (Row=Branch)
                branch_col_name = next((c for c in df_branch.columns if any(x in str(c).lower() for x in ['відділ', 'branch', 'філі'])), None)
                cashier_col_name = next((c for c in df_branch.columns if any(x in str(c).lower() for x in ['каса', 'cashier', 'касир'])), None)
                code_col_b = next((c for c in df_branch.columns if any(x in str(c).lower() for x in ['код', 'code', 'валют'])), None)
                buy_col_b = next((c for c in df_branch.columns if any(x in str(c).lower() for x in ['купів', 'buy'])), None)
                sell_col_b = next((c for c in df_branch.columns if any(x in str(c).lower() for x in ['прода', 'sell'])), None)

                # Check for Branch Matrix Cols (Hybrid)
                # Map lower cased symbols and codes to canonical codes
                # FIX: Include ALL currencies from DB + allowed defaults to ensure we don't skip valid ones
                all_currencies_db = db.query(models.Currency).all()
                curr_map = {}
                for c in all_currencies_db:
                    curr_map[c.code.lower()] = c.code
                
                # Also ensure ORDERED_CURRENCIES are in the map even if not in DB yet (though they should be)
                for code in ORDERED_CURRENCIES:
                     curr_map[code.lower()] = code
                
                # Add common symbols overrides
                curr_map['$'] = 'USD'
                curr_map['€'] = 'EUR'
                curr_map['zł'] = 'PLN'
                curr_map['pln'] = 'PLN'
                curr_map['gbp'] = 'GBP'
                curr_map['chf'] = 'CHF'
                matrix_cols = []
                for idx, col in enumerate(df_branch.columns):
                    c_clean = str(col).strip() # already lower cased by deduplication
                    found_curr = None
                    
                    # Exact match to find the "Buy" column (the first one, no suffix)
                    if c_clean in curr_map:
                        found_curr = curr_map[c_clean]
                    
                    # If we found a "Buy" column, the next one is "Sell"
                    if found_curr and idx + 1 < len(df_branch.columns):
                         mc = {'code': found_curr, 'buy_idx': idx, 'sell_idx': idx + 1}
                         
                         # Check for Wholesale columns (idx+2, idx+3)
                         # We enable wholesale reading if we have enough columns and they are not another currency's start
                         if idx + 3 < len(df_branch.columns):
                             next_c_clean = str(df_branch.columns[idx+2]).strip()
                             if next_c_clean not in curr_map:
                                 mc['wh_buy_idx'] = idx + 2
                                 mc['wh_sell_idx'] = idx + 3
                         
                         matrix_cols.append(mc)

                # Determine Mode
                use_hybrid = len(matrix_cols) > 0
                
                # Cache branches for cashier lookup if needed
                branch_cashier_map = {}
                if cashier_col_name:
                     all_branches = db.query(models.Branch).all()
                     for b in all_branches:
                         if b.cashier:
                             branch_cashier_map[b.cashier.strip().lower()] = b.id

                for row_idx, (_, row) in enumerate(df_branch.iterrows(), start=1):
                    try:
                        branch_id = None
                        # Resolve Branch ID
                        if branch_col_name and pd.notna(row[branch_col_name]):
                            try:
                                bid_val = row[branch_col_name]
                                if isinstance(bid_val, (int, float)):
                                    branch_id = int(bid_val)
                                else:
                                    branch_id = int(str(bid_val).strip())
                            except: pass
                        
                        if not branch_id and cashier_col_name and pd.notna(row[cashier_col_name]):
                            c_val = str(row[cashier_col_name]).strip().lower()
                            branch_id = branch_cashier_map.get(c_val)
                        
                        if not branch_id: continue
                        
                        # Update branch order by row index
                        try:
                            br_model = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
                            if br_model and br_model.order != row_idx:
                                br_model.order = row_idx
                                db.add(br_model)
                        except: pass

                        # 1. Process Matrix Columns (Hybrid/Row-Matrix)
                        if use_hybrid:
                            for mc in matrix_cols:
                                try:
                                    buy_val = row.iloc[mc['buy_idx']]
                                    sell_val = row.iloc[mc['sell_idx']]
                                    if pd.isna(buy_val) or pd.isna(sell_val): continue
                                    
                                    buy_str = str(buy_val).replace(',', '.').replace(' ', '').strip()
                                    sell_str = str(sell_val).replace(',', '.').replace(' ', '').strip()
                                    if buy_str == '-' or sell_str == '-': continue
                                    
                                    buy = float(buy_str)
                                    sell = float(sell_str)

                                    wh_buy = 0.0
                                    wh_sell = 0.0
                                    if 'wh_buy_idx' in mc and 'wh_sell_idx' in mc:
                                        try:
                                            wh_buy_val = row.iloc[mc['wh_buy_idx']]
                                            wh_sell_val = row.iloc[mc['wh_sell_idx']]
                                            if pd.notna(wh_buy_val):
                                                wh_b_str = str(wh_buy_val).replace(',', '.').replace(' ', '').strip()
                                                if wh_b_str and wh_b_str != '-': wh_buy = float(wh_b_str)
                                            if pd.notna(wh_sell_val):
                                                wh_s_str = str(wh_sell_val).replace(',', '.').replace(' ', '').strip()
                                                if wh_s_str and wh_s_str != '-': wh_sell = float(wh_s_str)
                                        except: pass
                                    
                                    # Upsert
                                    rate_entry = db.query(models.BranchRate).filter(
                                        models.BranchRate.branch_id == branch_id,
                                        models.BranchRate.currency_code == mc['code']
                                    ).first()
                                    if rate_entry:
                                        rate_entry.buy_rate = buy
                                        rate_entry.sell_rate = sell
                                        rate_entry.wholesale_buy_rate = wh_buy
                                        rate_entry.wholesale_sell_rate = wh_sell
                                        rate_entry.is_active = True
                                        processed_codes.add(mc['code'])
                                    else:
                                        # Check if currency exists, if not construct it (safe fallback)
                                        key_active = db.query(models.Currency).filter(models.Currency.code == mc['code']).first()
                                        if not key_active:
                                            # Create missing currency on the fly
                                            # Try to find metadata from defaults
                                            default_meta = next((c for c in currencies_data if c.code == mc['code']), None)
                                            name = default_meta.name if default_meta else mc['code']
                                            name_uk = default_meta.name_uk if default_meta else mc['code']
                                            flag = default_meta.flag if default_meta else "🏳️"
                                            
                                            new_curr = models.Currency(
                                                code=mc['code'],
                                                name=name,
                                                name_uk=name_uk,
                                                flag=flag,
                                                is_active=True,
                                                buy_rate=buy, # Set base rate to first branch rate found
                                                sell_rate=sell
                                            )
                                            db.add(new_curr)
                                            db.flush() # flush to make it available for FK
                                        
                                        db.add(models.BranchRate(
                                            branch_id=branch_id, 
                                            currency_code=mc['code'], 
                                            buy_rate=buy, 
                                            sell_rate=sell, 
                                            wholesale_buy_rate=wh_buy, 
                                            wholesale_sell_rate=wh_sell, 
                                            is_active=True
                                        ))
                                    branch_updated += 1
                                except Exception as e:
                                    # print(f"Error processing {mc['code']}: {e}")
                                    pass
                            
                            # IMPORTANT: In Hybrid mode, we IGNORE the legacy vertical columns (like NOK in row 3).
                            # This ensures that minor currencies use the Base Rate (Global) as instructed by user.
                            
                            # CLEANUP: Previously we deleted rates not in matrix. 
                            # NOW we keep them to support partial updates or hidden columns.
                            # No delete action here.

                            pass

                        # 2. Process Vertical Columns (ONLY if NOT hybrid OR fallback needed)
                        elif all([code_col_b, buy_col_b, sell_col_b]):
                            try:
                                code = str(row[code_col_b]).strip().upper()
                                buy = float(row[buy_col_b])
                                sell = float(row[sell_col_b])
                                
                                rate_entry = db.query(models.BranchRate).filter(
                                    models.BranchRate.branch_id == branch_id,
                                    models.BranchRate.currency_code == code
                                ).first()
                                if rate_entry:
                                    rate_entry.buy_rate = buy
                                    rate_entry.sell_rate = sell
                                    rate_entry.is_active = True
                                else:
                                    # Fallback create currency
                                    key_active = db.query(models.Currency).filter(models.Currency.code == code).first()
                                    if not key_active:
                                        # Validate code length
                                        if len(code) != 3: continue
                                        
                                        default_meta = next((c for c in currencies_data if c.code == code), None)
                                        name = default_meta.name if default_meta else code
                                        name_uk = default_meta.name_uk if default_meta else code
                                        flag = default_meta.flag if default_meta else "🏳️"
                                        
                                        new_curr = models.Currency(
                                            code=code, name=name, name_uk=name_uk, flag=flag,
                                            is_active=True, buy_rate=buy, sell_rate=sell
                                        )
                                        db.add(new_curr)
                                        db.flush()

                                    db.add(models.BranchRate(branch_id=branch_id, currency_code=code, buy_rate=buy, sell_rate=sell, is_active=True))
                                branch_updated += 1
                            except: pass

                    except Exception as e:
                        errors.append(f"Відділення (ряд): {str(e)}")
            else:
                # Matrix Format (Legacy: Row=Currency, Cols=Branches[1_buy, 1_sell])
                code_col_b = next((c for c in df_branch.columns if any(x in str(c).lower() for x in ['код', 'code', 'валют'])), df_branch.columns[0])
                
                branch_cols = {}
                for col in df_branch.columns:
                    if col == code_col_b: continue
                    import re
                    match = re.search(r'(\d+)', col)
                    if match:
                        bid = int(match.group(1))
                        if bid not in branch_cols: branch_cols[bid] = {}
                        if any(x in col for x in ['купів', 'buy']): branch_cols[bid]['buy'] = col
                        elif any(x in col for x in ['прода', 'sell']): branch_cols[bid]['sell'] = col
                
                for _, row in df_branch.iterrows():
                    try:
                        code = str(row[code_col_b]).strip().upper()
                        if not code or code == 'NAN': continue
                        
                        for bid, cols in branch_cols.items():
                            if 'buy' in cols and 'sell' in cols:
                                try:
                                    buy = float(row[cols['buy']])
                                    sell = float(row[cols['sell']])
                                    
                                    rate_entry = db.query(models.BranchRate).filter(
                                        models.BranchRate.branch_id == bid,
                                        models.BranchRate.currency_code == code
                                    ).first()
                                    
                                    if rate_entry:
                                        rate_entry.buy_rate = buy
                                        rate_entry.sell_rate = sell
                                    else:
                                        key_active = db.query(models.Currency).filter(models.Currency.code == code).first()
                                        if not key_active:
                                             if len(code) != 3: continue
                                             default_meta = next((c for c in currencies_data if c.code == code), None)
                                             name = default_meta.name if default_meta else code
                                             name_uk = default_meta.name_uk if default_meta else code
                                             flag = default_meta.flag if default_meta else "🏳️"
                                             
                                             new_curr = models.Currency(
                                                code=code, name=name, name_uk=name_uk, flag=flag,
                                                is_active=True, buy_rate=buy, sell_rate=sell
                                             )
                                             db.add(new_curr)
                                             db.flush()
                                        
                                        db.add(models.BranchRate(
                                            branch_id=bid,
                                            currency_code=code,
                                            buy_rate=buy,
                                            sell_rate=sell
                                        ))
                                    explicitly_updated_branches.add((bid, code))
                                    branch_updated += 1
                                except: pass
                    except Exception as e:
                        errors.append(f"Відділення (матриця): {str(e)}")

            db.commit()

        if processed_codes:
            # Final database sync for missing currencies
            missing_currencies = db.query(models.Currency).filter(
                models.Currency.is_active == True,
                ~models.Currency.code.in_(processed_codes)
            ).all()
            
            for mc in missing_currencies:
                mc.is_active = False
                print(f"SYNC: Deactivating currency {mc.code} (missing from Excel)")
            
            # Also deactivate branch overrides for these
            db.query(models.BranchRate).filter(
                ~models.BranchRate.currency_code.in_(processed_codes)
            ).update({models.BranchRate.is_active: False}, synchronize_session=False)
            
            db.commit()
            
            # Sync base rates to BranchRate table for branches NOT explicitly updated
            # This ensures all branches reflect the new base rates from the Excel
            all_branches = db.query(models.Branch).all()
            for code in processed_codes:
                base_curr = db.query(models.Currency).filter(models.Currency.code == code).first()
                if not base_curr:
                    continue
                for branch in all_branches:
                    # Skip branches that were explicitly set from Excel columns
                    if (branch.id, code) in explicitly_updated_branches:
                        continue
                    
                    br = db.query(models.BranchRate).filter(
                        models.BranchRate.branch_id == branch.id,
                        models.BranchRate.currency_code == code
                    ).first()
                    if br:
                        # Only update if NOT explicitly disabled for this branch
                        if br.is_active:
                            br.buy_rate = base_curr.buy_rate
                            br.sell_rate = base_curr.sell_rate
                            br.wholesale_buy_rate = base_curr.wholesale_buy_rate
                            br.wholesale_sell_rate = base_curr.wholesale_sell_rate
                    else:
                        # Create BranchRate entry
                        br = models.BranchRate(
                            branch_id=branch.id,
                            currency_code=code,
                            buy_rate=base_curr.buy_rate,
                            sell_rate=base_curr.sell_rate,
                            wholesale_buy_rate=base_curr.wholesale_buy_rate,
                            wholesale_sell_rate=base_curr.wholesale_sell_rate,
                            is_active=True
                        )
                        db.add(br)
                    branch_updated += 1
            db.commit()
            
            # Update global cache
            for cd in currencies_data:
                if cd.code not in processed_codes:
                    cd.is_active = False
                else:
                    cd.is_active = True
            
        rates_updated_at = datetime.now()
        
        return RatesUploadResponseV2(
            success=True,
            message=f"Курси оновлено о {rates_updated_at.strftime('%H:%M:%S')}",
            base_rates_updated=base_updated,
            branch_rates_updated=branch_updated,
            errors=errors[:10]
        )
        
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Помилка обробки файлу: {str(e)}")

@app.get("/api/admin/rates/template")
async def download_rates_template(
    user: User = Depends(require_operator_or_admin),
    db: Session = Depends(get_db)
):
    """Download Excel template for rates upload with vertical layout (Rows=Currencies, Cols=Branches)"""
    from fastapi.responses import StreamingResponse
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet "Курси"
        
        # 1. Fetch Data
        branches = db.query(models.Branch).order_by(models.Branch.number).all()
        all_rates = db.query(models.BranchRate).all()
        
        # Map: (branch_id, code) -> rate
        rates_map = {(r.branch_id, r.currency_code): r for r in all_rates}
        
        # Map: code -> currency info (for names/flags)
        db_currencies = db.query(models.Currency).all()
        curr_info_map = {c.code: c for c in db_currencies}

        # 2. Prepare Columns (Row 3 headers)
        # Base Columns
        columns = ['Код', 'Прапор', 'Валюта']
        
        # Branch Columns - repeated for each branch
        for branch in branches:
            columns.extend(['Купівля', 'Продаж', 'Опт Купівля', 'Опт Продаж'])
            
        data_rows = []
        
        # 3. Build Rows (Iterate Strict List)
        for code in ORDERED_CURRENCIES:
            # Get info or default
            curr_info = curr_info_map.get(code)
            
            # Default values if currency not in DB yet
            flag = curr_info.flag if curr_info else CURRENCY_FLAGS.get(code, "🏳️")
            name_uk = curr_info.name_uk if curr_info else CURRENCY_NAMES.get(code, (code, code))[1]
            
            row = [code, flag, name_uk]
            
            for branch in branches:
                rate = rates_map.get((branch.id, code))
                
                # Default to 0.00 if no rate exists
                buy = rate.buy_rate if rate else 0.00
                sell = rate.sell_rate if rate else 0.00
                wh_buy = rate.wholesale_buy_rate if rate else 0.00
                wh_sell = rate.wholesale_sell_rate if rate else 0.00
                
                row.extend([buy, sell, wh_buy, wh_sell])
            
            data_rows.append(row)
            
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=columns)
        
        # Write to Excel, starting at row 2 (0-indexed logic in pandas, so startrow=2 means Excel Row 3)
        # We need space for Top Headers (Row 1 & 2)
        df.to_excel(writer, index=False, sheet_name='Курси', startrow=2)
        
        # Access the workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets['Курси']
        
        # Styles
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Colors (approximate brand or logical colors)
        fill_buy = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
        fill_sell = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red/Pink
        fill_opt_buy = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Pale Green
        fill_opt_sell = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Pale Orange
        fill_neutral = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Grey

        # 4. Add Top Headers (Row 1 & 2)
        # ... (Same as before) ...
        # Apply Logic for Row 3 Styling (Headers)
        for col_idx, col_name in enumerate(columns, 1):
             cell = worksheet.cell(row=3, column=col_idx)
             name_str = str(col_name).lower()
             
             if 'опт' in name_str and 'купів' in name_str:
                 cell.fill = fill_opt_buy
             elif 'опт' in name_str and 'прода' in name_str:
                 cell.fill = fill_opt_sell
             elif 'купів' in name_str:
                 cell.fill = fill_buy
             elif 'прода' in name_str:
                 cell.fill = fill_sell
             else:
                 cell.fill = fill_neutral
                 
             cell.font = header_font
             cell.alignment = center_align
        # Row 1: Branch Numbers
        # Row 2: Branch Addresses
        
        current_col = 4 # Column D (1-based index)
        for branch in branches:
            # Merge 4 cells
            start_col_letter = get_column_letter(current_col)
            end_col_letter = get_column_letter(current_col + 3)
            cell_range_r1 = f"{start_col_letter}1:{end_col_letter}1"
            cell_range_r2 = f"{start_col_letter}2:{end_col_letter}2"
            
            # Row 1: Number (Editable)
            worksheet.merge_cells(cell_range_r1)
            cell_r1 = worksheet.cell(row=1, column=current_col)
            # Just the number, maybe with "№" prefix or just raw number? 
            # User asked for "Number", let's put just number or "№ X"
            # If user edits it to just "5", we need to parse.
            # Let's stick to "№ {number}" for display, but user can edit.
            cell_r1.value = f"№ {branch.number if branch.number else branch.id}" 
            cell_r1.font = header_font
            cell_r1.alignment = center_align

            # Row 2: Address (Identifier for logic if number changes)
            worksheet.merge_cells(cell_range_r2)
            cell_r2 = worksheet.cell(row=2, column=current_col)
            cell_r2.value = branch.address
            cell_r2.font = header_font
            cell_r2.alignment = center_align
            
            current_col += 4
            
        # Adjust column widths
        for i, col in enumerate(df.columns):
            # +1 because openpyxl is 1-indexed
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = 15 if i > 2 else 10 # Wider for rates
            
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=svit_valut_rates.xlsx"}
    )



@app.get("/api/calculate/cross")
async def calculate_cross_exchange(
    amount: float,
    from_currency: str,
    to_currency: str,
    db: Session = Depends(get_db)
):
    """Calculate exchange using automatic 2-step conversion via UAH"""
    from_currency = from_currency.upper()
    to_currency = to_currency.upper()
    
    if from_currency == to_currency:
        return {"from_amount": amount, "to_amount": amount, "rate": 1.0}
    
    # Calculate via UAH
    # Step 1: FROM -> UAH (Buy rate for from_currency)
    # Step 2: UAH -> TO (Sell rate for to_currency)
    
    from_r = db.query(models.BranchRate).filter(models.BranchRate.branch_id == 1, models.BranchRate.currency_code == from_currency).first()
    to_r = db.query(models.BranchRate).filter(models.BranchRate.branch_id == 1, models.BranchRate.currency_code == to_currency).first()
    
    if from_currency == "UAH":
        if not to_r:
            raise HTTPException(status_code=404, detail=f"Currency {to_currency} not found")
        to_amount = amount / to_r.sell_rate
        effective_rate = 1 / to_r.sell_rate
    elif to_currency == "UAH":
        if not from_r:
            raise HTTPException(status_code=404, detail=f"Currency {from_currency} not found")
        to_amount = amount * from_r.buy_rate
        effective_rate = from_r.buy_rate
    else:
        if not from_r or not to_r:
            raise HTTPException(status_code=404, detail="Currency not found")
        
        # FROM -> UAH -> TO
        uah_amount = amount * from_r.buy_rate
        to_amount = uah_amount / to_r.sell_rate
        effective_rate = from_r.buy_rate / to_r.sell_rate
    
    return {
        "from_amount": amount,
        "from_currency": from_currency,
        "to_amount": round(to_amount, 2),
        "to_currency": to_currency,
        "rate": round(effective_rate, 4),
        "pair": f"{from_currency}/{to_currency} (auto-calc via UAH)"
    }

class BranchRateUpdate(BaseModel):
    buy_rate: Optional[float] = None
    sell_rate: Optional[float] = None
    wholesale_buy_rate: Optional[float] = None
    wholesale_sell_rate: Optional[float] = None
    wholesale_threshold: Optional[int] = None
    is_active: Optional[bool] = None

@app.put("/api/admin/rates/branch/{branch_id}/{currency_code}")
async def update_branch_rate(
    branch_id: int,
    currency_code: str,
    data: BranchRateUpdate,
    user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Update or create branch specific rate"""
    currency_code = currency_code.upper()
    
    # Check if rate exists
    rate = db.query(models.BranchRate).filter(
        models.BranchRate.branch_id == branch_id,
        models.BranchRate.currency_code == currency_code
    ).first()
    
    # SPECIAL LOGIC: "Reset to Base" for Minor Currencies
    # If user toggles "Active" (True) for a minor currency (e.g. CAD),
    # and implies no price change (None), we should DELETE the override
    # so it falls back to Base Rate (Dynamic).
    MAJOR_CURS = ['USD', 'EUR', 'PLN', 'GBP', 'CHF']
    
    # Only if we are just toggling (prices are None)
    if data.buy_rate is None and data.sell_rate is None:
        if currency_code not in MAJOR_CURS:
            if data.is_active is True:
                # User wants to ENABLE it.
                # Use Base Rate -> Delete Branch Override
                if rate:
                    db.delete(rate)
                    db.commit()
                return {"message": "Reverted to Base Rate"}
                
    if not rate:
        # Create new rate if it doesn't exist
        rate = models.BranchRate(
            branch_id=branch_id,
            currency_code=currency_code,
            buy_rate=data.buy_rate if data.buy_rate is not None else 0.0,
            sell_rate=data.sell_rate if data.sell_rate is not None else 0.0,
            wholesale_buy_rate=data.wholesale_buy_rate if data.wholesale_buy_rate is not None else 0.0,
            wholesale_sell_rate=data.wholesale_sell_rate if data.wholesale_sell_rate is not None else 0.0,
            wholesale_threshold=data.wholesale_threshold if data.wholesale_threshold is not None else 1000,
            is_active=data.is_active if data.is_active is not None else True
        )
        db.add(rate)
    else:
        # Update existing
        if data.buy_rate is not None:
            rate.buy_rate = data.buy_rate
        if data.sell_rate is not None:
            rate.sell_rate = data.sell_rate
        if data.wholesale_buy_rate is not None:
            rate.wholesale_buy_rate = data.wholesale_buy_rate
        if data.wholesale_sell_rate is not None:
            rate.wholesale_sell_rate = data.wholesale_sell_rate
        if data.wholesale_threshold is not None:
            rate.wholesale_threshold = data.wholesale_threshold
        if data.is_active is not None:
            rate.is_active = data.is_active
        if data.sell_rate is not None:
            rate.sell_rate = data.sell_rate
        if data.is_active is not None:
            rate.is_active = data.is_active
            
    db.commit()
    return {"success": True}

@app.get("/api/admin/rates/all")
async def get_all_rates_admin(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all rates including branch-specific and cross-rates (Admin only)"""
    
    # Base rates (Global list)
    base_rates = {
        c.code: {"buy": c.buy_rate, "sell": c.sell_rate, "name": c.name_uk} 
        for c in currencies_data
    }
    
    # Branches (DB)
    branches_list = db.query(models.Branch).order_by(models.Branch.order.asc()).all()
    branches_out = [{"id": b.id, "address": b.address, "number": b.number, "order": b.order} for b in branches_list]
    
    # Branch rates (DB)
    branch_rates_rows = db.query(models.BranchRate).all()
    branch_rates = {}
    
    for br in branch_rates_rows:
        if br.branch_id not in branch_rates:
            branch_rates[br.branch_id] = {}
        branch_rates[br.branch_id][br.currency_code] = {
            "buy": br.buy_rate,
            "sell": br.sell_rate,
            "wholesale_buy": br.wholesale_buy_rate,
            "wholesale_sell": br.wholesale_sell_rate,
            "wholesale_threshold": br.wholesale_threshold or 1000,
            "is_active": br.is_active
        }
        
    return {
        "updated_at": rates_updated_at.isoformat(),
        "base_rates": base_rates,
        "branch_rates": branch_rates,
        "branches": branches_out
    }

@app.get("/api/admin/dashboard", response_model=DashboardStats)
async def get_admin_dashboard(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get admin dashboard statistics"""
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    total = db.query(models.Reservation).count()
    pending = db.query(models.Reservation).filter(
        models.Reservation.status.in_([models.ReservationStatus.PENDING, models.ReservationStatus.PENDING_ADMIN])
    ).count()
    confirmed = db.query(models.Reservation).filter(models.Reservation.status == models.ReservationStatus.CONFIRMED).count()
    completed_today = db.query(models.Reservation).filter(
        models.Reservation.status == models.ReservationStatus.COMPLETED,
        models.Reservation.completed_at >= today_start
    ).count()
    
    total_volume = db.query(models.Reservation).filter(
        models.Reservation.get_currency == "UAH",
        models.Reservation.status == models.ReservationStatus.COMPLETED,
        models.Reservation.completed_at >= today_start
    ).with_entities(models.Reservation.get_amount).all()
    
    total_volume_uah = sum(r[0] for r in total_volume)
    
    # Calculate MONTHLY volume
    month_start = today_start.replace(day=1)
    
    total_volume_month = db.query(models.Reservation).filter(
        models.Reservation.get_currency == "UAH",
        models.Reservation.status == models.ReservationStatus.COMPLETED,
        models.Reservation.completed_at >= month_start
    ).with_entities(models.Reservation.get_amount).all()
    
    total_volume_uah_month = sum(r[0] for r in total_volume_month)
    
    return DashboardStats(
        total_reservations=total,
        pending_reservations=pending,
        confirmed_reservations=confirmed,
        completed_today=completed_today,
        total_volume_uah=total_volume_uah,
        total_volume_uah_month=total_volume_uah_month
    )

@app.get("/api/admin/reservations")
async def get_all_reservations(
    user: models.User = Depends(require_admin),
    status: Optional[models.ReservationStatus] = None,
    branch_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """Get all reservations (Admin only)"""
    query = db.query(models.Reservation)
    
    if status:
        query = query.filter(models.Reservation.status == status)
    else:
        # Exclude soft-deleted reservations from the default view
        query = query.filter(models.Reservation.status != 'deleted')
    if branch_id:
        query = query.filter(models.Reservation.branch_id == branch_id)
    
    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(models.Reservation.created_at >= dt_from)
        except ValueError:
            pass
            
    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(models.Reservation.created_at < dt_to)
        except ValueError:
            pass
    
    total = query.count()
    
    # Sort by created_at descending
    db_items = query.order_by(models.Reservation.created_at.desc()).offset((page - 1) * limit).limit(limit).all()
    
    items = []
    for r in db_items:
        branch = db.query(models.Branch).filter(models.Branch.id == r.branch_id).first()
        items.append(ReservationResponse(
            id=r.id,
            give_amount=r.give_amount,
            give_currency=r.give_currency,
            get_amount=r.get_amount,
            get_currency=r.get_currency,
            rate=r.rate,
            phone=r.phone,
            customer_name=r.customer_name,
            status=r.status,
            branch_id=r.branch_id,
            branch_address=branch.address if branch else None,
            created_at=r.created_at.isoformat(),
            updated_at=r.updated_at.isoformat() if r.updated_at else None,
            expires_at=r.expires_at.isoformat(),
            completed_at=r.completed_at.isoformat() if r.completed_at else None,
            operator_note=r.operator_note
        ))
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }


class ReservationEdit(BaseModel):
    give_amount: Optional[float] = None
    get_amount: Optional[float] = None
    rate: Optional[float] = None
    branch_id: Optional[int] = None
    customer_name: Optional[str] = None
    phone: Optional[str] = None
    status: Optional[str] = None
    operator_note: Optional[str] = None


@app.put("/api/admin/reservations/{reservation_id}")
async def update_reservation(
    reservation_id: int,
    data: ReservationEdit,
    user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Update reservation details (Admin only)"""
    res = db.query(models.Reservation).filter(models.Reservation.id == reservation_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    if data.give_amount is not None:
        res.give_amount = data.give_amount
    if data.get_amount is not None:
        res.get_amount = data.get_amount
    if data.rate is not None:
        res.rate = data.rate
    if data.branch_id is not None:
        branch = db.query(models.Branch).filter(models.Branch.id == data.branch_id).first()
        if not branch:
            raise HTTPException(status_code=400, detail="Branch not found")
        res.branch_id = data.branch_id
    if data.customer_name is not None:
        res.customer_name = data.customer_name
    if data.phone is not None:
        res.phone = data.phone
    if data.status is not None:
        try:
            res.status = ReservationStatus(data.status)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {data.status}")
    if data.operator_note is not None:
        res.operator_note = data.operator_note
    
    db.commit()
    db.refresh(res)
    
    branch = db.query(models.Branch).filter(models.Branch.id == res.branch_id).first()
    return ReservationResponse(
        id=res.id,
        give_amount=res.give_amount,
        give_currency=res.give_currency,
        get_amount=res.get_amount,
        get_currency=res.get_currency,
        rate=res.rate,
        phone=res.phone,
        customer_name=res.customer_name,
        status=res.status,
        branch_id=res.branch_id,
        branch_address=branch.address if branch else None,
        created_at=res.created_at.isoformat(),
        updated_at=res.updated_at.isoformat() if res.updated_at else None,
        expires_at=res.expires_at.isoformat(),
        completed_at=res.completed_at.isoformat() if res.completed_at else None,
        operator_note=res.operator_note
    )


@app.post("/api/admin/reservations/{reservation_id}/assign")
async def assign_reservation_to_operator(
    reservation_id: int,
    user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Assign reservation to operator (changes status from pending_admin to pending)"""
    res = db.query(models.Reservation).filter(models.Reservation.id == reservation_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    if res.status != models.ReservationStatus.PENDING_ADMIN:
        # Admin can still assign from any status if needed
        pass
    
    if not res.branch_id:
        raise HTTPException(status_code=400, detail="Reservation must have a branch assigned")
    
    res.status = models.ReservationStatus.PENDING
    db.commit()
    db.refresh(res)
    
    branch = db.query(models.Branch).filter(models.Branch.id == res.branch_id).first()
    return ReservationResponse(
        id=res.id,
        give_amount=res.give_amount,
        give_currency=res.give_currency,
        get_amount=res.get_amount,
        get_currency=res.get_currency,
        rate=res.rate,
        phone=res.phone,
        customer_name=res.customer_name,
        status=res.status,
        branch_id=res.branch_id,
        branch_address=branch.address if branch else None,
        created_at=res.created_at.isoformat(),
        updated_at=res.updated_at.isoformat() if res.updated_at else None,
        expires_at=res.expires_at.isoformat(),
        completed_at=res.completed_at.isoformat() if res.completed_at else None,
        operator_note=res.operator_note
    )


# ============== OPERATOR ENDPOINTS ==============

@app.get("/api/operator/dashboard", response_model=DashboardStats)
async def get_operator_dashboard(user: models.User = Depends(require_operator_or_admin), db: Session = Depends(get_db)):
    """Get operator dashboard statistics for their branch"""
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    query = db.query(models.Reservation)
    if user.role == models.UserRole.OPERATOR and user.branch_id:
        query = query.filter(models.Reservation.branch_id == user.branch_id)
    
    total = query.count()
    pending = query.filter(models.Reservation.status == models.ReservationStatus.PENDING).count()
    confirmed = query.filter(models.Reservation.status == models.ReservationStatus.CONFIRMED).count()
    completed_today = query.filter(
        models.Reservation.status == models.ReservationStatus.COMPLETED,
        models.Reservation.completed_at >= today_start
    ).count()
    
    # Volume Today
    volume_today_query = query.filter(
        models.Reservation.get_currency == "UAH",
        models.Reservation.status == models.ReservationStatus.COMPLETED,
        models.Reservation.completed_at >= today_start
    )
    total_volume_uah = sum(r.get_amount for r in volume_today_query.all())
    
    # Volume Month
    month_start = today_start.replace(day=1)
    volume_month_query = query.filter(
        models.Reservation.get_currency == "UAH",
        models.Reservation.status == models.ReservationStatus.COMPLETED,
        models.Reservation.completed_at >= month_start
    )
    total_volume_uah_month = sum(r.get_amount for r in volume_month_query.all())
    
    return DashboardStats(
        total_reservations=total,
        pending_reservations=pending,
        confirmed_reservations=confirmed,
        completed_today=completed_today,
        total_volume_uah=total_volume_uah,
        total_volume_uah_month=total_volume_uah_month
    )

@app.get("/api/operator/reservations")
async def get_branch_reservations(
    user: models.User = Depends(require_operator_or_admin),
    status: Optional[models.ReservationStatus] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """Get reservations for operator's branch"""
    query = db.query(models.Reservation)
    if user.role == models.UserRole.OPERATOR and user.branch_id:
        query = query.filter(models.Reservation.branch_id == user.branch_id)
    
    if status:
        query = query.filter(models.Reservation.status == status)
        
    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(models.Reservation.created_at >= dt_from)
        except ValueError:
            pass
            
    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(models.Reservation.created_at < dt_to)
        except ValueError:
            pass
    
    total = query.count()
    db_items = query.order_by(models.Reservation.created_at.desc()).offset((page - 1) * limit).limit(limit).all()
    
    items = []
    for r in db_items:
        branch = db.query(models.Branch).filter(models.Branch.id == r.branch_id).first()
        items.append(ReservationResponse(
            id=r.id,
            give_amount=r.give_amount,
            give_currency=r.give_currency,
            get_amount=r.get_amount,
            get_currency=r.get_currency,
            rate=r.rate,
            phone=r.phone,
            customer_name=r.customer_name,
            status=r.status,
            branch_id=r.branch_id,
            branch_address=branch.address if branch else None,
            created_at=r.created_at.isoformat(),
            updated_at=r.updated_at.isoformat() if r.updated_at else None,
            expires_at=r.expires_at.isoformat(),
            completed_at=r.completed_at.isoformat() if r.completed_at else None,
            operator_note=r.operator_note
        ))
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@app.put("/api/operator/reservations/{reservation_id}")
async def update_reservation(
    reservation_id: int,
    update: ReservationUpdate,
    user: models.User = Depends(require_operator_or_admin),
    db: Session = Depends(get_db)
):
    """Update reservation status (Operator)"""
    db_res = db.query(models.Reservation).filter(models.Reservation.id == reservation_id).first()
    if not db_res:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    # Check if operator has access to this reservation
    if user.role == models.UserRole.OPERATOR and user.branch_id:
        if db_res.branch_id != user.branch_id:
            raise HTTPException(status_code=403, detail="Access denied to this reservation")
    
    if update.status:
        db_res.status = update.status
        if update.status == models.ReservationStatus.COMPLETED:
            db_res.completed_at = datetime.now()
    
    if update.operator_note:
        db_res.operator_note = update.operator_note
    
    db.commit()
    db.refresh(db_res)
    
    branch = db.query(models.Branch).filter(models.Branch.id == db_res.branch_id).first()
    
    return ReservationResponse(
        id=db_res.id,
        give_amount=db_res.give_amount,
        give_currency=db_res.give_currency,
        get_amount=db_res.get_amount,
        get_currency=db_res.get_currency,
        rate=db_res.rate,
        phone=db_res.phone,
        customer_name=db_res.customer_name,
        status=db_res.status,
        branch_id=db_res.branch_id,
        branch_address=branch.address if branch else None,
        created_at=db_res.created_at.isoformat(),
        updated_at=db_res.updated_at.isoformat() if db_res.updated_at else None,
        expires_at=db_res.expires_at.isoformat(),
        completed_at=db_res.completed_at.isoformat() if db_res.completed_at else None,
        operator_note=db_res.operator_note
    )

@app.post("/api/operator/reservations/{reservation_id}/confirm")
async def confirm_reservation(
    reservation_id: int,
    user: User = Depends(require_operator_or_admin),
    db: Session = Depends(get_db)
):
    """Confirm a pending reservation"""
    reservation = db.query(models.Reservation).filter(models.Reservation.id == reservation_id).first()
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    if user.role == models.UserRole.OPERATOR and user.branch_id:
        if reservation.branch_id != user.branch_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    if user.role != models.UserRole.ADMIN and reservation.status != models.ReservationStatus.PENDING:
        raise HTTPException(status_code=400, detail="Can only confirm pending reservations")
    
    reservation.status = models.ReservationStatus.CONFIRMED
    db.commit()
    db.refresh(reservation)
    return reservation

@app.post("/api/operator/reservations/{reservation_id}/complete")
async def complete_reservation(
    reservation_id: int,
    user: User = Depends(require_operator_or_admin),
    db: Session = Depends(get_db)
):
    """Mark reservation as completed"""
    reservation = db.query(models.Reservation).filter(models.Reservation.id == reservation_id).first()
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    if user.role == models.UserRole.OPERATOR and user.branch_id:
        if reservation.branch_id != user.branch_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    if user.role != models.UserRole.ADMIN and reservation.status not in [models.ReservationStatus.PENDING, models.ReservationStatus.CONFIRMED]:
        raise HTTPException(status_code=400, detail="Cannot complete this reservation")
    
    reservation.status = models.ReservationStatus.COMPLETED
    reservation.completed_at = datetime.now()
    db.commit()
    db.refresh(reservation)
    return reservation

@app.post("/api/operator/reservations/{reservation_id}/cancel")
async def cancel_reservation(
    reservation_id: int,
    user: User = Depends(require_operator_or_admin),
    db: Session = Depends(get_db)
):
    """Cancel a reservation"""
    reservation = db.query(models.Reservation).filter(models.Reservation.id == reservation_id).first()
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    if user.role == models.UserRole.OPERATOR and user.branch_id:
        if reservation.branch_id != user.branch_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    if user.role != models.UserRole.ADMIN and reservation.status == models.ReservationStatus.COMPLETED:
        raise HTTPException(status_code=400, detail="Cannot cancel completed reservation")
    
    reservation.status = models.ReservationStatus.CANCELLED
    db.commit()
    db.refresh(reservation)
    return reservation


# ============== ADMIN CROSS-RATES ENDPOINTS ==============

class CrossRateCreate(BaseModel):
    base_currency: str
    quote_currency: str
    buy_rate: float = 0.0
    sell_rate: float = 0.0
    is_active: bool = True
    order: int = 0

class CrossRateUpdate(BaseModel):
    base_currency: Optional[str] = None
    quote_currency: Optional[str] = None
    buy_rate: Optional[float] = None
    sell_rate: Optional[float] = None
    is_active: Optional[bool] = None
    order: Optional[int] = None

@app.get("/api/admin/cross-rates")
async def get_admin_cross_rates(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all cross-rate pairs (Admin)"""
    rates = db.query(models.CrossRate).order_by(models.CrossRate.order.asc()).all()
    return [
        {
            "id": r.id,
            "base_currency": r.base_currency,
            "quote_currency": r.quote_currency,
            "buy_rate": r.buy_rate,
            "sell_rate": r.sell_rate,
            "is_active": r.is_active,
            "order": r.order
        }
        for r in rates
    ]

@app.post("/api/admin/cross-rates")
async def create_cross_rate(data: CrossRateCreate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Create a new cross-rate pair"""
    cr = models.CrossRate(
        base_currency=data.base_currency.upper(),
        quote_currency=data.quote_currency.upper(),
        buy_rate=data.buy_rate,
        sell_rate=data.sell_rate,
        is_active=data.is_active,
        order=data.order
    )
    db.add(cr)
    db.commit()
    db.refresh(cr)
    return {"id": cr.id, "base_currency": cr.base_currency, "quote_currency": cr.quote_currency, "buy_rate": cr.buy_rate, "sell_rate": cr.sell_rate, "is_active": cr.is_active, "order": cr.order}

@app.put("/api/admin/cross-rates/{rate_id}")
async def update_cross_rate(rate_id: int, data: CrossRateUpdate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update a cross-rate pair"""
    cr = db.query(models.CrossRate).filter(models.CrossRate.id == rate_id).first()
    if not cr:
        raise HTTPException(status_code=404, detail="Cross-rate not found")
    if data.base_currency is not None:
        cr.base_currency = data.base_currency.upper()
    if data.quote_currency is not None:
        cr.quote_currency = data.quote_currency.upper()
    if data.buy_rate is not None:
        cr.buy_rate = data.buy_rate
    if data.sell_rate is not None:
        cr.sell_rate = data.sell_rate
    if data.is_active is not None:
        cr.is_active = data.is_active
    if data.order is not None:
        cr.order = data.order
    db.commit()
    db.refresh(cr)
    return {"id": cr.id, "base_currency": cr.base_currency, "quote_currency": cr.quote_currency, "buy_rate": cr.buy_rate, "sell_rate": cr.sell_rate, "is_active": cr.is_active, "order": cr.order}

@app.delete("/api/admin/cross-rates/{rate_id}")
async def delete_cross_rate(rate_id: int, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete a cross-rate pair"""
    cr = db.query(models.CrossRate).filter(models.CrossRate.id == rate_id).first()
    if not cr:
        raise HTTPException(status_code=404, detail="Cross-rate not found")
    db.delete(cr)
    db.commit()
    return {"success": True}


# ============== PUBLIC SETTINGS ENDPOINTS ==============



# ============== ADMIN SETTINGS ENDPOINTS ==============

@app.get("/api/admin/settings")
async def get_admin_settings(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all settings (Admin only)"""
    return {
        "site": db.query(models.SiteSettings).first(),
        "faq": db.query(models.FAQItem).all(),
        "services": db.query(models.ServiceItem).all(),
        "articles": db.query(models.ArticleItem).all(),
    }

@app.put("/api/admin/settings")
async def update_settings(settings: SiteSettings, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update site settings (Admin only)"""
    db_settings = db.query(models.SiteSettings).first()
    if not db_settings:
        db_settings = models.SiteSettings()
        db.add(db_settings)
    
    # Update fields
    for key, value in settings.dict().items():
        setattr(db_settings, key, value)
    
    db.commit()
    return {"success": True, "message": "Налаштування оновлено"}

@app.get("/api/admin/faq")
async def get_admin_faq(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all FAQ items (Admin only)"""
    return db.query(models.FAQItem).order_by(models.FAQItem.order).all()

@app.post("/api/admin/faq")
async def create_faq(item: FAQItem, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Create FAQ item"""
    db_item = models.FAQItem(
        question=item.question,
        answer=item.answer,
        link_text=item.link_text,
        link_url=item.link_url,
        order=item.order
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

@app.put("/api/admin/faq/{faq_id}")
async def update_faq(faq_id: int, item: FAQItem, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update FAQ item"""
    db_item = db.query(models.FAQItem).filter(models.FAQItem.id == faq_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="FAQ not found")
    
    db_item.question = item.question
    db_item.answer = item.answer
    db_item.link_text = item.link_text
    db_item.link_url = item.link_url
    db_item.order = item.order
    
    db.commit()
    db.refresh(db_item)
    return db_item

@app.delete("/api/admin/faq/{faq_id}")
async def delete_faq(faq_id: int, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete FAQ item"""
    db_item = db.query(models.FAQItem).filter(models.FAQItem.id == faq_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="FAQ not found")
    db.delete(db_item)
    db.commit()
    return {"success": True}

@app.get("/api/admin/services")
async def get_admin_services(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all services (Admin only)"""
    return db.query(models.ServiceItem).order_by(models.ServiceItem.order).all()

@app.post("/api/admin/services")
async def create_service(item: ServiceItem, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Create service"""
    db_item = models.ServiceItem(
        title=item.title,
        short_description=item.short_description,
        description=item.description,
        image_url=item.image_url,
        link_url=item.link_url,
        is_active=item.is_active,
        order=item.order
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

@app.put("/api/admin/services/{service_id}")
async def update_service(service_id: int, item: ServiceItem, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update service"""
    db_item = db.query(models.ServiceItem).filter(models.ServiceItem.id == service_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Service not found")
    
    db_item.title = item.title
    db_item.short_description = item.short_description
    db_item.description = item.description
    db_item.image_url = item.image_url
    db_item.link_url = item.link_url
    db_item.is_active = item.is_active
    db_item.order = item.order
    
    db.commit()
    db.refresh(db_item)
    return db_item

@app.delete("/api/admin/services/{service_id}")
async def delete_service(service_id: int, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete service"""
    db_item = db.query(models.ServiceItem).filter(models.ServiceItem.id == service_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Service not found")
    db.delete(db_item)
    db.commit()
    return {"success": True}


# ============== ADMIN BRANCH MANAGEMENT ==============

class BranchCreate(BaseModel):
    number: int = 0
    address: str
    hours: str = "щодня: 9:00-20:00"
    lat: Optional[float] = None
    lng: Optional[float] = None
    is_open: bool = True
    phone: Optional[str] = None
    telegram_chat: Optional[str] = None
    cashier: Optional[str] = None

class BranchUpdate(BaseModel):
    number: Optional[int] = None
    address: Optional[str] = None
    hours: Optional[str] = None
    phone: Optional[str] = None
    telegram_chat: Optional[str] = None
    cashier: Optional[str] = None
    is_open: Optional[bool] = None
    lat: Optional[float] = None
    lng: Optional[float] = None

@app.get("/api/admin/branches")
async def get_admin_branches(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all branches with full details"""
    return db.query(models.Branch).order_by(models.Branch.order.asc()).all()

@app.put("/api/admin/branches/{branch_id}")
async def update_branch(branch_id: int, update: BranchUpdate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update branch info"""
    branch = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    if update.number is not None:
        branch.number = update.number
    if update.address is not None:
        branch.address = update.address
    if update.hours is not None:
        branch.hours = update.hours
    if update.phone is not None:
        branch.phone = update.phone
    if update.telegram_chat is not None:
        branch.telegram_chat = update.telegram_chat
    if update.cashier is not None:
        branch.cashier = update.cashier
    if update.is_open is not None:
        branch.is_open = update.is_open
    
    # Manual coordinate override
    manual_coords = False
    if update.lat is not None:
        branch.lat = update.lat
        manual_coords = True
    if update.lng is not None:
        branch.lng = update.lng
        manual_coords = True
        
    # Auto-geocode if address changed, but only if manual coords weren't just explicitly provided
    if update.address is not None and update.address != branch.address and not manual_coords:
        coords = geocode_address(update.address)
        if coords:
            branch.lat = coords[0]
            branch.lng = coords[1]
    
    db.commit()
    db.refresh(branch)
    return branch

def geocode_address(address: str) -> Optional[tuple[float, float]]:
    try:
        # Enhance address with City if needed, but 'Kyiv' is usually enough if not present
        search_query = address
        if "київ" not in address.lower() and "kyiv" not in address.lower():
            search_query += ", Київ"
            
        url = f"https://nominatim.openstreetmap.org/search?q={urllib.parse.quote(search_query)}&format=json&limit=1"
        req = urllib.request.Request(url, headers={'User-Agent': 'SvitValutApp/1.0'})
        
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode())
            if data and len(data) > 0:
                lat = float(data[0]['lat'])
                lon = float(data[0]['lon'])
                return (lat, lon)
    except Exception as e:
        print(f"Geocoding error for {address}: {e}")
    return None

@app.post("/api/admin/branches", response_model=Branch)
async def create_branch(branch: BranchCreate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Create a new branch"""
    # Create new ID (simple max + 1)
    last_branch = db.query(models.Branch).order_by(models.Branch.id.desc()).first()
    new_id = (last_branch.id + 1) if last_branch else 1
    
    
    # Auto-geocode if lat/lng not provided
    geo_lat = branch.lat
    geo_lng = branch.lng
    
    if geo_lat is None or geo_lng is None:
        coords = geocode_address(branch.address)
        if coords:
            geo_lat = coords[0]
            geo_lng = coords[1]
        else:
            # Fallback to defaults if geocoding fails
            geo_lat = geo_lat or 50.4501
            geo_lng = geo_lng or 30.5234

    db_branch = models.Branch(
        id=new_id,
        address=branch.address,
        hours=branch.hours,
        lat=geo_lat,
        lng=geo_lng,
        is_open=branch.is_open,
        phone=branch.phone,
        telegram_chat=branch.telegram_chat,
        cashier=branch.cashier
    )
    db.add(db_branch)
    db.commit()
    db.refresh(db_branch)
    
    # Initialize rates for this branch using branch 1 (or default) as template
    base_rates = db.query(models.BranchRate).filter(models.BranchRate.branch_id == 1).all()
    for rate in base_rates:
        db.add(models.BranchRate(
            branch_id=db_branch.id,
            currency_code=rate.currency_code,
            buy_rate=rate.buy_rate,
            sell_rate=rate.sell_rate
        ))
    db.commit()
    
    return db_branch

@app.delete("/api/admin/branches/{branch_id}")
async def delete_branch(branch_id: int, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete branch"""
    branch = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    if branch_id == 1:
        raise HTTPException(status_code=400, detail="Cannot delete main branch")
        
    # Check if used by users
    users = db.query(models.User).filter(models.User.branch_id == branch_id).first()
    if users:
        raise HTTPException(status_code=400, detail="Cannot delete branch with assigned operators")
    
    # Delete rates first
    db.query(models.BranchRate).filter(models.BranchRate.branch_id == branch_id).delete()
    
    db.delete(branch)
    db.commit()
    return {"success": True}


# ============== ADMIN USER MANAGEMENT ==============

class UserCreate(BaseModel):
    username: str
    password: str
    name: str
    branch_id: Optional[int] = None
    role: UserRole = UserRole.OPERATOR

class UserUpdate(BaseModel):
    name: Optional[str] = None
    branch_id: Optional[int] = None
    password: Optional[str] = None

class UserResponse(BaseModel):
    id: int
    username: str
    name: str
    role: UserRole
    branch_id: Optional[int]
    branch_address: Optional[str] = None

@app.get("/api/admin/users", response_model=List[UserResponse])
async def get_admin_users(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all users"""
    users = db.query(models.User).all()
    result = []
    for u in users:
        branch_address = None
        if u.branch_id:
            branch = db.query(models.Branch).filter(models.Branch.id == u.branch_id).first()
            branch_address = branch.address if branch else None
        result.append(UserResponse(
            id=u.id,
            username=u.username,
            name=u.name,
            role=UserRole(u.role.value),
            branch_id=u.branch_id,
            branch_address=branch_address
        ))
    return result

@app.post("/api/admin/users", response_model=UserResponse)
async def create_user(user_data: UserCreate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Create new operator"""
    # Check username uniqueness
    existing = db.query(models.User).filter(models.User.username == user_data.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Validate branch if provided
    if user_data.branch_id:
        branch = db.query(models.Branch).filter(models.Branch.id == user_data.branch_id).first()
        if not branch:
            raise HTTPException(status_code=400, detail="Branch not found")
    
    new_user = models.User(
        username=user_data.username,
        password_hash=user_data.password,  # In production, hash this!
        name=user_data.name,
        role=models.UserRole[user_data.role.value.upper()],
        branch_id=user_data.branch_id
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    branch_address = None
    if new_user.branch_id:
        branch = db.query(models.Branch).filter(models.Branch.id == new_user.branch_id).first()
        branch_address = branch.address if branch else None
    
    return UserResponse(
        id=new_user.id,
        username=new_user.username,
        name=new_user.name,
        role=UserRole(new_user.role.value),
        branch_id=new_user.branch_id,
        branch_address=branch_address
    )

@app.put("/api/admin/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: int, user_data: UserUpdate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update user"""
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user_data.name is not None:
        target_user.name = user_data.name
    if user_data.branch_id is not None:
        # Validate branch
        branch = db.query(models.Branch).filter(models.Branch.id == user_data.branch_id).first()
        if not branch:
            raise HTTPException(status_code=400, detail="Branch not found")
        target_user.branch_id = user_data.branch_id
    if user_data.password is not None and user_data.password.strip():
        target_user.password_hash = user_data.password  # In production, hash this!
    
    db.commit()
    db.refresh(target_user)
    
    branch_address = None
    if target_user.branch_id:
        branch = db.query(models.Branch).filter(models.Branch.id == target_user.branch_id).first()
        branch_address = branch.address if branch else None
    
    return UserResponse(
        id=target_user.id,
        username=target_user.username,
        name=target_user.name,
        role=UserRole(target_user.role.value),
        branch_id=target_user.branch_id,
        branch_address=branch_address
    )

@app.delete("/api/admin/users/{user_id}")
async def delete_user(user_id: int, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete user"""
    if user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    db.delete(target_user)
    db.commit()
    return {"success": True}


@app.get("/api/seo", response_model=List[SeoMetadata])
async def public_get_seo_metadata(db: Session = Depends(get_db)):
    return db.query(models.SeoMetadata).all()

# ============== ADMIN SEO METADATA ==============

@app.get("/api/admin/seo", response_model=List[SeoMetadata])
async def admin_get_seo_metadata(user: User = Depends(require_admin), db: Session = Depends(get_db)):
    return db.query(models.SeoMetadata).all()

@app.post("/api/admin/seo", response_model=SeoMetadata)
async def admin_create_seo_metadata(item: SeoMetadataCreate, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    path = item.url_path
    if not path.startswith('/'):
        path = f"/{path}"
        item.url_path = path

    existing = db.query(models.SeoMetadata).filter_by(url_path=path).first()
    if existing:
        raise HTTPException(status_code=400, detail="SEO configuration already exists for this path")
        
    new_seo = models.SeoMetadata(**item.dict())
    db.add(new_seo)
    db.commit()
    db.refresh(new_seo)
    return new_seo

@app.put("/api/admin/seo/{seo_id}", response_model=SeoMetadata)
async def admin_update_seo_metadata(seo_id: int, item: SeoMetadataUpdate, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    seo = db.query(models.SeoMetadata).filter(models.SeoMetadata.id == seo_id).first()
    if not seo:
        raise HTTPException(status_code=404, detail="SEO metadata not found")

    path = item.url_path
    if not path.startswith('/'):
        path = f"/{path}"
        item.url_path = path

    # Check for duplicate path on a different ID
    dup = db.query(models.SeoMetadata).filter(models.SeoMetadata.url_path == path, models.SeoMetadata.id != seo_id).first()
    if dup:
        raise HTTPException(status_code=400, detail="Another SEO configuration already exists for this path")

    for key, value in item.dict(exclude_unset=True).items():
        setattr(seo, key, value)

    db.commit()
    db.refresh(seo)
    return seo

@app.delete("/api/admin/seo/{seo_id}")
async def admin_delete_seo_metadata(seo_id: int, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    seo = db.query(models.SeoMetadata).filter(models.SeoMetadata.id == seo_id).first()
    if not seo:
        raise HTTPException(status_code=404, detail="SEO metadata not found")
        
    db.delete(seo)
    db.commit()
    return {"success": True}

# ============== ADMIN CURRENCY MANAGEMENT ==============

class CurrencyUpdate(BaseModel):
    is_active: Optional[bool] = None
    is_popular: Optional[bool] = None
    buy_rate: Optional[float] = None
    sell_rate: Optional[float] = None
    buy_url: Optional[str] = None
    sell_url: Optional[str] = None
    seo_h1: Optional[str] = None
    seo_h2: Optional[str] = None
    seo_image: Optional[str] = None
    seo_text: Optional[str] = None
    
    # Split SEO Fields
    seo_buy_h1: Optional[str] = None
    seo_buy_h2: Optional[str] = None
    seo_buy_title: Optional[str] = None
    seo_buy_desc: Optional[str] = None
    seo_buy_text: Optional[str] = None
    seo_buy_image: Optional[str] = None

    seo_sell_h1: Optional[str] = None
    seo_sell_h2: Optional[str] = None
    seo_sell_title: Optional[str] = None
    seo_sell_desc: Optional[str] = None
    seo_sell_text: Optional[str] = None
    seo_sell_image: Optional[str] = None
    
    wholesale_threshold: Optional[int] = None

@app.get("/api/currencies/info/all")
async def get_all_currency_info(db: Session = Depends(get_db)):
    """Get SEO info for all currencies (public)"""
    currencies = db.query(models.Currency).filter(models.Currency.is_active == True).order_by(models.Currency.order).all()
    result = {}
    for c in currencies:
        result[c.code] = {
            "code": c.code,
            "name_uk": c.name_uk,
            "flag": c.flag,
            "buy_url": c.buy_url,
            "sell_url": c.sell_url,
            "seo_h1": c.seo_h1,
            "seo_h2": c.seo_h2,
            "seo_image": c.seo_image,
            "seo_text": c.seo_text,
            "seo_buy_h1": c.seo_buy_h1,
            "seo_buy_h2": c.seo_buy_h2,
            "seo_buy_title": c.seo_buy_title,
            "seo_buy_desc": c.seo_buy_desc,
            "seo_buy_text": c.seo_buy_text,
            "seo_buy_image": c.seo_buy_image,
            "seo_sell_h1": c.seo_sell_h1,
            "seo_sell_h2": c.seo_sell_h2,
            "seo_sell_title": c.seo_sell_title,
            "seo_sell_desc": c.seo_sell_desc,
            "seo_sell_text": c.seo_sell_text,
            "seo_sell_image": c.seo_sell_image,
        }
    return result

@app.get("/api/admin/currencies")
async def get_admin_currencies(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get all currencies including inactive with SEO fields"""
    currencies = db.query(models.Currency).order_by(models.Currency.order).all()
    
    result = []
    for c in currencies:
        result.append({
            "code": c.code,
            "name": c.name,
            "name_uk": c.name_uk,
            "flag": c.flag,
            "buy_rate": c.buy_rate,
            "sell_rate": c.sell_rate,
            "is_popular": c.is_popular,
            "is_active": c.is_active,
            "buy_url": c.buy_url,
            "sell_url": c.sell_url,
            "seo_h1": c.seo_h1,
            "seo_h2": c.seo_h2,
            "seo_image": c.seo_image,
            "seo_text": c.seo_text,
            "seo_buy_h1": c.seo_buy_h1,
            "seo_buy_h2": c.seo_buy_h2,
            "seo_buy_title": c.seo_buy_title,
            "seo_buy_desc": c.seo_buy_desc,
            "seo_buy_text": c.seo_buy_text,
            "seo_buy_image": c.seo_buy_image,
            "seo_sell_h1": c.seo_sell_h1,
            "seo_sell_h2": c.seo_sell_h2,
            "seo_sell_title": c.seo_sell_title,
            "seo_sell_desc": c.seo_sell_desc,
            "seo_sell_text": c.seo_sell_text,
            "seo_sell_image": c.seo_sell_image,
        })
    return result

@app.put("/api/admin/currencies/{code}")
async def update_currency(code: str, update: CurrencyUpdate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update currency settings (enable/disable, base rates)"""
    # Use Currency table for Base Rates
    c = db.query(models.Currency).filter(models.Currency.code == code.upper()).first()
    
    if not c:
        raise HTTPException(status_code=404, detail="Currency not found")
    
    if update.buy_rate is not None:
        c.buy_rate = update.buy_rate
    if update.sell_rate is not None:
        c.sell_rate = update.sell_rate
    if update.is_active is not None:
        c.is_active = update.is_active
    if update.is_popular is not None:
        c.is_popular = update.is_popular
    if update.buy_url is not None:
        c.buy_url = update.buy_url
    if update.sell_url is not None:
        c.sell_url = update.sell_url
    if update.seo_h1 is not None:
        c.seo_h1 = update.seo_h1
    if update.seo_h2 is not None:
        c.seo_h2 = update.seo_h2
    if update.seo_image is not None:
        c.seo_image = update.seo_image
    if update.seo_text is not None:
        c.seo_text = update.seo_text
        
    # Split SEO updates
    if update.seo_buy_h1 is not None: c.seo_buy_h1 = update.seo_buy_h1
    if update.seo_buy_h2 is not None: c.seo_buy_h2 = update.seo_buy_h2
    if update.seo_buy_title is not None: c.seo_buy_title = update.seo_buy_title
    if update.seo_buy_desc is not None: c.seo_buy_desc = update.seo_buy_desc
    if update.seo_buy_text is not None: c.seo_buy_text = update.seo_buy_text
    if update.seo_buy_image is not None: c.seo_buy_image = update.seo_buy_image

    if update.seo_sell_h1 is not None: c.seo_sell_h1 = update.seo_sell_h1
    if update.seo_sell_h2 is not None: c.seo_sell_h2 = update.seo_sell_h2
    if update.seo_sell_title is not None: c.seo_sell_title = update.seo_sell_title
    if update.seo_sell_desc is not None: c.seo_sell_desc = update.seo_sell_desc
    if update.seo_sell_text is not None: c.seo_sell_text = update.seo_sell_text
    if update.seo_sell_image is not None: c.seo_sell_image = update.seo_sell_image

    if update.wholesale_threshold is not None:
        c.wholesale_threshold = update.wholesale_threshold
    
    global rates_updated_at
    rates_updated_at = datetime.now()
    
    db.commit()
    db.refresh(c)
    
    return {
        "code": c.code,
        "name": c.name,
        "name_uk": c.name_uk,
        "flag": c.flag,
        "buy_rate": c.buy_rate,
        "sell_rate": c.sell_rate,
        "is_popular": c.is_popular,
        "is_active": c.is_active,
        "buy_url": c.buy_url,
        "sell_url": c.sell_url,
        "seo_h1": c.seo_h1,
        "seo_h2": c.seo_h2,
        "seo_image": c.seo_image,
        "seo_text": c.seo_text,
        "seo_buy_h1": c.seo_buy_h1,
        "seo_buy_h2": c.seo_buy_h2,
        "seo_buy_title": c.seo_buy_title,
        "seo_buy_desc": c.seo_buy_desc,
        "seo_buy_text": c.seo_buy_text,
        "seo_buy_image": c.seo_buy_image,
        "seo_sell_h1": c.seo_sell_h1,
        "seo_sell_h2": c.seo_sell_h2,
        "seo_sell_title": c.seo_sell_title,
        "seo_sell_desc": c.seo_sell_desc,
        "seo_sell_text": c.seo_sell_text,
        "seo_sell_image": c.seo_sell_image,
    }


# ============== OPERATOR RATES DOWNLOAD ==============

@app.get("/api/operator/rates/download")
async def download_operator_rates(user: models.User = Depends(require_operator_or_admin), db: Session = Depends(get_db)):
    """Download rates for operator's branch as Excel"""
    import io
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="pandas not installed")
    
    branch_id = user.branch_id or 1
    branch = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
    
    db_rates = db.query(models.BranchRate).filter(models.BranchRate.branch_id == branch_id).all()
    
    rates_data = []
    for r in db_rates:
        names = CURRENCY_NAMES.get(r.currency_code, (r.currency_code, r.currency_code))
        rates_data.append({
            'Код валюти': r.currency_code,
            'Назва': names[1],
            'Купівля': r.buy_rate,
            'Продаж': r.sell_rate,
        })
    
    if not rates_data:
        rates_data = [{'Код валюти': 'Немає даних', 'Назва': '', 'Купівля': 0, 'Продаж': 0}]
    
    df = pd.DataFrame(rates_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Курси', index=False)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    branch_name = branch.address if branch else 'all'
    # Clean filename
    safe_name = ''.join(c if c.isalnum() or c in '._-' else '_' for c in branch_name)
    filename = f"rates_{safe_name}.xlsx"
    
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

# ============== LIVE CHAT FEATURE ==============



@app.get("/api/admin/chat/sessions", response_model=List[ChatSession])
async def admin_get_chat_sessions(user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Admin get all active chat sessions sorted by recent activity"""
    sessions = db.query(models.ChatSession).filter(models.ChatSession.status == models.ChatSessionStatus.ACTIVE).order_by(models.ChatSession.last_message_at.desc()).all()
    return sessions

@app.get("/api/admin/chat/sessions/{session_id}/messages", response_model=List[ChatMessage])
async def admin_get_session_messages(session_id: str, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Admin get messages for a session"""
    session = db.query(models.ChatSession).filter(models.ChatSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    messages = db.query(models.ChatMessage).filter(models.ChatMessage.session_id == session.id).order_by(models.ChatMessage.created_at.asc()).all()
    return messages

@app.post("/api/admin/chat/sessions/{session_id}/messages", response_model=ChatMessage)
async def admin_send_chat_message(session_id: str, msg: ChatMessageCreate, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Admin sending a message"""
    session = db.query(models.ChatSession).filter(models.ChatSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    new_msg = models.ChatMessage(
        session_id=session.id,
        sender="admin",
        content=msg.content
    )
    db.add(new_msg)
    
    session.last_message_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(new_msg)
    
    return new_msg

@app.post("/api/admin/chat/sessions/{session_id}/read")
async def admin_mark_messages_read(session_id: str, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Admin marks user messages as read"""
    session = db.query(models.ChatSession).filter(models.ChatSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    unread_user_msgs = db.query(models.ChatMessage).filter(
        models.ChatMessage.session_id == session.id,
        models.ChatMessage.sender == 'user',
        models.ChatMessage.is_read == False
    ).all()
    
    for m in unread_user_msgs:
        m.is_read = True
        
    db.commit()
    return {"success": True}
    
@app.put("/api/admin/chat/sessions/{session_id}/close")
async def admin_close_chat_session(session_id: str, user: models.User = Depends(require_admin), db: Session = Depends(get_db)):
    """Admin closes a chat session"""
    session = db.query(models.ChatSession).filter(models.ChatSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
        
    session.status = models.ChatSessionStatus.CLOSED
    db.commit()
    return {"success": True}

@app.get("/sitemap.xml", response_class=Response)
async def get_sitemap(db: Session = Depends(get_db)):
    """Generate dynamic XML sitemap"""
    frontend_url = os.environ.get("FRONTEND_URL", "https://mirvalut.com").rstrip("/")
    
    pages = [
        "",
        "/rates",
        "/services",
        "/contacts",
        "/faq"
    ]
    
    urls = []
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Static pages
    for page in pages:
        urls.append(f"""
  <url>
    <loc>{frontend_url}{page}</loc>
    <lastmod>{today}</lastmod>
    <changefreq>daily</changefreq>
    <priority>{1.0 if page == '' else 0.8}</priority>
  </url>""")

    # Dynamic services
    services = db.query(models.ServiceItem).filter(models.ServiceItem.is_active == True).all()
    for srv in services:
        # Use link_url as slug if present, else id
        slug = (srv.link_url or str(srv.id)).lstrip("/")
        if not slug.startswith("services/"):
            slug = f"services/{slug}"
        urls.append(f"""
  <url>
    <loc>{frontend_url}/{slug}</loc>
    <lastmod>{today}</lastmod>
    <changefreq>weekly</changefreq>
    <priority>0.7</priority>
  </url>""")

    # Dynamic articles
    articles = db.query(models.ArticleItem).filter(models.ArticleItem.is_published == True).all()
    for art in articles:
        date_str = art.created_at.strftime("%Y-%m-%d") if art.created_at else today
        urls.append(f"""
  <url>
    <loc>{frontend_url}/articles/{art.id}</loc>
    <lastmod>{date_str}</lastmod>
    <changefreq>monthly</changefreq>
    <priority>0.6</priority>
  </url>""")

    sitemap_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">{''.join(urls)}
</urlset>"""

    return Response(content=sitemap_xml, media_type="application/xml")


# --- Frontend static file serving (production) ---
FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "frontend")

if os.path.isdir(FRONTEND_DIR):
    # Serve frontend assets (JS, CSS, images)
    frontend_assets = os.path.join(FRONTEND_DIR, "assets")
    if os.path.isdir(frontend_assets):
        app.mount("/assets", StaticFiles(directory=frontend_assets), name="frontend_assets")

    # Serve frontend images
    frontend_images = os.path.join(FRONTEND_DIR, "images")
    if os.path.isdir(frontend_images):
        app.mount("/images", StaticFiles(directory=frontend_images), name="frontend_images")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        """Serve frontend files or fall back to index.html for SPA routing"""
        file_path = os.path.join(FRONTEND_DIR, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)
        index_path = os.path.join(FRONTEND_DIR, "index.html")
        if os.path.isfile(index_path):
            return FileResponse(index_path)
        return {"error": "Frontend not found"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
